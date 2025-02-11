# See LICENSE file for copyright and license details.
"""
Requiments order utils.

Functions
---------
- To export to excel

"""
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

from follow_control_card.models import Card


ITEM_COLUMN = 6
DAY_SPANISH = {
    'Monday': 'Lunes',
    'Tuesday': 'Martes',
    'Wednesday': 'Miércoles',
    'Thursday': 'Jueves',
    'Friday': 'Viernes',
    'Saturday': 'Sábado',
    'Sunday': 'Domingo',
}


def card_excel_prepare(user):
    """
    Make a format to user card by month and add basic date.
    """

    # New excel book
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Calendario de {user.username}"

    # Get cards
    today = datetime.today()

    # Encabezado con detalles de la orden
    ws['A1'] = "Datos"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:K1')

    details = [
        f"Mes: {today.month}",
        f"Usuario: {user.username}",
    ]

    for idx, detail in enumerate(details, start=2):
        ws[f"A{idx}"] = detail

    return wb, ws


def card_excel_header(ws):
    """
    Add the headers to card excel.
    """

    headers = [
        "Dia del mes",
        "Dia",
        "Orden de venta (SAP)",
        "Orden de venta (Descripción)",
        "Verbo",
        "Objeto",
        "Unidad de medida",
        "Tiempo",
    ]
    header_fill = PatternFill(
        start_color="FFC000",
        end_color="FFC000",
        fill_type="solid"
    )
    header_font = Font(bold=True)
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=ITEM_COLUMN, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Change the column's width
    column_widths = [
        15,
        20,
        25,
        30,
        10,
        20,
        15,
        25,
        25,
        15,
        10
    ]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    return ws


def card_excel_item(wb, ws, user):
    """
    Add items to card excel.
    """

    today = datetime.today()
    cards = Card.objects.filter(
        user=user,
        date__year=today.year,
        date__month=today.month
    )

    print("-----")
    print(cards)
    print("-----")

    # Add item's data
    # TODO: add process and subprocess
    row_num = ITEM_COLUMN
    for card in cards:
        tasks = card.tasks.order_by('cardtaskorder__order')

        print("--taks---")
        print(tasks)
        print("-----")

        for task in tasks:
            print(">>>>")
            print(task)
            data = [
                str(card.date.day),
                DAY_SPANISH[card.date.strftime('%A')],
                str(task.sale_order.sapcode) or "N/A",
                task.sale_order.detail or "N/A",
                task.verb or "N/A",
                task.object or "N/A",
                task.measurement or "N/A",
                task.task_time or "N/A",
            ]
            row_num += 1
            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                # cell.border = border
                # if col_num == 10:  # Estado
                #    cell.alignment = Alignment(horizontal="center")
                # else:
                #    cell.alignment = Alignment(horizontal="left")
                cell.alignment = Alignment(horizontal="left")

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    )
    response["Content-Disposition"] = 'attachment; filename="mis_tareas.xlsx"'
    wb.save(response)

    return response
