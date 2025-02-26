# See LICENSE file for copyright and license details.
"""
Request view
"""
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import DetailView
from django.forms import inlineformset_factory
from django.contrib import messages
from django.views.decorators.http import require_POST
from django.core.exceptions import ValidationError
from django.core.cache import cache

from accounting_order_sales.models import SalesOrder, SalesOrderItem
from logistic_requirements.models import (
    RequirementOrder,
    RequirementOrderItem)
from logistic_suppliers.models import Suppliers
from .forms import (
    CreateRequirementOrderForm,
    CreateRequirementOrderItemFormSet,
    PrepopulatedRequirementOrderItemForm
    )


def index_requests(request):
    """
    Main view of order requests
    """
    # pylint: disable=no-member
    sales_orders = SalesOrder.objects.filter(is_active=True).order_by('-id')
    context = {'sales_orders': sales_orders}
    return render(request, 'index_requests.html', context)


def requirement_orders_view(request, sales_order_id):
    """
    view to see the requirement orders for each order
    """
    sales_order = get_object_or_404(SalesOrder, id=sales_order_id)

    requirement_orders = sales_order.requirement_orders.all().order_by('requested_date')

    context = {
        'sales_order': sales_order,
        'requirement_orders': requirement_orders,
    }

    return render(request, 'requirement_orders_list.html', context)

def my_requests(request):
    """
    Views orders filtered by user

    Filtered
    --------
    user : admin
        admin orders only
    """
    # pylint: disable=no-member
    my_orders = RequirementOrder.objects.filter(
        user=request.user).order_by('-id')
    context = {'my_orders': my_orders}
    return render(request, 'requests/my_requests.html', context)


def total_requests(request):
    """
    View of all orders
    """
    # pylint: disable=no-member
    total_orders = RequirementOrder.objects.all().order_by('-id')
    context = {'total_orders': total_orders}

    return render(request, 'total_requests/all_requests.html', context)


def requirement_order_preview(request):
    """
    Launches a modal that shows the status of the items
    It is a part of a larger template, like a partials
    """
    order_id = request.GET.get('order_id')
    requirement_order = get_object_or_404(RequirementOrder, id=order_id)
    items = requirement_order.items.all()  # Obtener los ítems asociados
    context = {'order': requirement_order, 'items': items}
    return render(request, 'requests/requirement_order_preview.html', context)


def delete_order(request, order_id):
    """Delete an order"""
    if request.method == 'DELETE':
        order = get_object_or_404(
            RequirementOrder, id=order_id, user=request.user)
        order.delete()
        return HttpResponse(
            '<div class="alert alert-success">¡' +
            'El pedido ha sido eliminado correctamente!</div>',
            content_type='text/html')
    return HttpResponse(
        '<div class="alert alert-danger">' +
        'Hubo un problema al eliminar el pedido.</div>',
        status=400)


def create_requests(request, order_id):
    """Create a new order"""
    sales_order = get_object_or_404(SalesOrder, id=order_id)
    # pylint: disable=no-member
    referencia_ordenventa = SalesOrderItem.objects.filter(
        salesorder=sales_order)

    if request.method == "POST":
        order_form = CreateRequirementOrderForm(request.POST)
        formset = CreateRequirementOrderItemFormSet(
            request.POST,
            request.FILES,
            form_kwargs={'sales_order': sales_order})

        if order_form.is_valid() and formset.is_valid():
            try:
                requirement_order = order_form.save(commit=False)
                requirement_order.sales_order = sales_order
                requirement_order.user = request.user        # Diccionario para almacenar las cantidades solicitadas por ítem
                # Diccionario para almacenar las
                # cantidades solicitadas por ítem
                item_quantities = {}
                valid_items = []

                for form in formset:
                    item = form.save(commit=False)

                    # Asegurarse de que el sales_order_item
                    # está correctamente asignado
                    if not item.sales_order_item:
                        raise ValidationError(
                            f"El ítem '{item.description}' " +
                            "no tiene un artículo de orden de venta asociado.")

                    # Obtener el ID del sales_order_item
                    sales_order_item_id = item.sales_order_item.id

                    # Verificar la cantidad total solicitada,
                    # sumando las cantidades repetidas
                    if sales_order_item_id in item_quantities:
                        total_quantity = (
                            item_quantities[sales_order_item_id] +
                            item.quantity_requested
                        )
                    else:
                        total_quantity = item.quantity_requested

                    # Verificar que la cantidad total solicitada
                    # no exceda la cantidad disponible
                    if total_quantity > item.sales_order_item.remaining_requirement:
                        raise ValidationError(
                            f"La cantidad solicitada ({total_quantity}) "
                            f"para '{item.sales_order_item.description}' "
                            f"excede la cantidad disponible "
                            f"({item.sales_order_item.remaining_requirement}).")
                    # Actualizar el diccionario con la cantidad solicitada
                    item_quantities[sales_order_item_id] = total_quantity

                    # Asegurarse de que el precio se obtiene correctamente
                    item.price = item.price or item.sales_order_item.price
                    # Validación de cantidad
                    if item.sales_order_item.remaining_requirement <= 0 and item.quantity_requested > 0:
                        raise ValidationError(
                            f"El ítem '{item.sales_order_item.description}' "
                            f"tiene cantidad cero disponible "
                            f"y no puede ser solicitado.")

                    item.clean()  # Limpiar y validar el ítem
                    valid_items.append(item)  # Añadir el ítem válido a la lista

                if not valid_items:
                    raise ValidationError(
                        "No hay ítems válidos para agregar al pedido.")

                # Guardar la orden si hay al menos un ítem válido
                requirement_order.save()
                for item in valid_items:
                    item.requirement_order = requirement_order
                    item.save()

                messages.success(request, "Pedido creado exitosamente.")
                return redirect('index_requests')

            except ValidationError as e:
                messages.error(request, str(e))

        else:
            for form in formset:
                if form.errors:
                    item_name = form.cleaned_data.get(
                        'sales_order_item',
                        'Sin nombre')
                    for field, errors in form.errors.items():
                        for error in errors:
                            messages.error(
                                request,
                                f"Error en {item_name}: {field} - {error}")
            messages.error(
                request,
                "Por favor, corrija los errores en el formulario.")
    
    else:
        order_form = CreateRequirementOrderForm(
            initial={'sales_order': sales_order})
        formset = CreateRequirementOrderItemFormSet(
            queryset=RequirementOrderItem.objects.none(),
            form_kwargs={'sales_order': sales_order}
        )
    context = {
        'order_form': order_form,
        'formset': formset,
        'sales_order': sales_order,
        'referencia_ordenventa': referencia_ordenventa, }
    return render(request, 'requests/create_requests.html', context)


def create_prepopulated_request(request, order_id):
    """
    Create a new requirement order with prepopulated data
    """
    sales_order = get_object_or_404(SalesOrder, id=order_id)
    # pylint: disable=no-member
    sales_order_items = SalesOrderItem.objects.filter(
        salesorder=sales_order)

    # Prepare initial data
    initial_items = [
        {
            'sales_order_item': item.id,
            'quantity_requested': item.amount,
        }
        for item in sales_order_items
    ]

    prepopulated_formset = inlineformset_factory(
        RequirementOrder,
        RequirementOrderItem,
        form=PrepopulatedRequirementOrderItemForm,
        extra=len(initial_items),
        can_delete=True  # Allow users to delete items
    )

    if request.method == "POST":
        order_form = CreateRequirementOrderForm(request.POST)
        formset = prepopulated_formset(
            request.POST,
            form_kwargs={'sales_order': sales_order}
        )
        print(formset.is_valid())
        if order_form.is_valid() and formset.is_valid():
            try:
                requirement_order = order_form.save(commit=False)
                requirement_order.sales_order = sales_order
                requirement_order.user = request.user
                requirement_order.save()

                items = formset.save(commit=False)
                for item in items:
                    item.requirement_order = requirement_order
                    item.save()

                messages.success(request, "Pedido creado exitosamente.")
                return redirect('index_requests')

            except ValidationError as e:
                messages.error(request, str(e))
        else:
            # Display errors for debugging
            print("Order form errors:", order_form.errors)
            print("Formset errors:", formset.errors)
            messages.error(
                request,
                "Por favor, corrija los errores en el formulario.")
    else:
        order_form = CreateRequirementOrderForm(
            initial={'sales_order': sales_order})
        formset = prepopulated_formset(
            queryset=RequirementOrderItem.objects.none(),
            initial=initial_items,
            form_kwargs={'sales_order': sales_order}
        )

    # Attach item descriptions to each form
    for form, item in zip(formset.forms, sales_order_items):
        form.item_description = item.description

    # Crear lista combinada para pasar a la plantilla
    combined_data = [
        {
            'form': form,
            'unit_of_measurement': item.unit_of_measurement,
            'price': item.price,
            'price_total': item.price_total,
        }
        for form, item in zip(formset.forms, sales_order_items)
    ]
    context = {
        'order_form': order_form,
        'formset': formset,
        'sales_order': sales_order,
        'combined_data': combined_data,  # Lista combinada para la plantilla
    }
    return render(
        request,
        'requests/create_prepopulated_request.html',
        context)


class MyRequestDetail(DetailView):
    """ View to see the detail of a request """
    model = RequirementOrder
    template_name = 'requests/my_request_detail.html'
    context_object_name = 'order'


def delete_requirement_order_item(request, item_id):
    """Function to delete an item from a requirment order"""
    # Obtener el item a eliminar y su RequirementOrder asociado
    item = get_object_or_404(RequirementOrderItem, id=item_id)
    order = item.requirement_order
    # Eliminar el item
    item.delete()
    # Re-renderizar el partial `items_table.html`
    # con la lista de ítems actualizada
    context = {'order': order}
    return render(request, 'partials/item_table.html', context)


def ajax_load_suppliers(request):
    """View to load suppliers via AJAX"""
    term = request.GET.get(
        'term', '').strip().lower()  # Filtrar y limpiar el término de búsqueda
    # Intenta recuperar de caché si ya se ha buscado este término
    cache_key = f"suppliers_search_{term}"
    supplier_list = cache.get(cache_key)

    if not supplier_list:  # Si no está en caché, realiza la consulta
        # pylint: disable=no-member
        suppliers = Suppliers.objects.filter(
            name__icontains=term
            ).only('id', 'name')[:20]
        supplier_list = [
            {
                'id': supplier.id,
                'text': supplier.name} for supplier in suppliers]
        cache.set(
            cache_key,
            supplier_list,
            timeout=60 * 5)  # Cachea la respuesta por 5 minutos

    return JsonResponse({'results': supplier_list})


# Nueva vista para pedidos rapidos
def request_sales_order(request, pk):
    """View to create a requirement order from a sales order"""
    # Usamos select_related para evitar consultas adicionales de clave externa
    # pylint: disable=no-member
    sales_order = get_object_or_404(
        SalesOrder.objects.select_related('project'), id=pk)
    # Prefetch_related permite cargar `requirementorderitem_set`
    # de forma eficiente
    sales_order_items = sales_order.items.all().prefetch_related(
        'requirementorderitem_set'
        )
    # Elimina la carga de suppliers
    context = {
        'sales_order': sales_order,
        'sales_order_items': sales_order_items,
    }
    return render(request, 'requests_plus/requests_plus.html', context)


@require_POST
def create_requirement_order(request, order_id):
    """Create a requirement order from a sales order"""
    try:
        sales_order = get_object_or_404(SalesOrder, id=order_id)
        order_has_errors = False
        items_to_create = []

        # Mostrar el contenido de request.POST para depurar
        print("[Debug] request.POST items:", request.POST.items())

        for item_id, quantity_requested in request.POST.items():
            if item_id.startswith("items-") and "-quantity_requested" in item_id:
                sales_order_item_id = int(item_id.split("-")[1])
                sales_order_item = get_object_or_404(
                    SalesOrderItem,
                    id=sales_order_item_id)

                try:
                    quantity_requested = float(quantity_requested)
                except ValueError as ve:
                    print(
                        f"[Error] Cantidad inválida en item_id {item_id}: {ve}"
                        )
                    messages.error(
                        request,
                        f"Cantidad solicitada inválida para el ítem"
                        f" {sales_order_item.description}. "
                        f"Debe ser un número.")
                    order_has_errors = True
                    continue

                price = request.POST.get(
                    f"items-{sales_order_item_id}-price",
                    sales_order_item.price
                    )
                try:
                    price = float(price)
                except ValueError as ve:
                    print(
                        f"[Error] Precio inválido "
                        f"en item_id {item_id}: {ve}")
                    messages.error(
                        request,
                        f"Precio inválido para el ítem"
                        f"{sales_order_item.description}.")
                    order_has_errors = True
                    continue

                # Verificar que la cantidad solicitada no
                # exceda el remaining_requirement
                if quantity_requested > sales_order_item.remaining_requirement:
                    print(
                        f"[Error] La cantidad solicitada ({quantity_requested})"
                        f"para '{sales_order_item.description}' excede la "
                        f"cantidad disponible ({sales_order_item.remaining_requirement}).")
                    messages.error(
                        request,
                        f"La cantidad solicitada para '{sales_order_item.description}'"
                        f" excede la cantidad disponible.")
                    order_has_errors = True
                    continue

                # Verificar que el remaining_requirement sea mayor que cero
                if sales_order_item.remaining_requirement <= 0:
                    print(f"[Error] No hay cantidad disponible para"
                          f" '{sales_order_item.description}'.")
                    messages.error(
                        request,
                        f"No hay cantidad disponible para el ítem"
                        f" '{sales_order_item.description}'.")
                    order_has_errors = True
                    continue

                # Solo agregar ítem si tiene cantidad solicitada mayor que cero
                if quantity_requested > 0:
                    supplier_id = request.POST.get(
                        f"items-{sales_order_item_id}-supplier")
                    notes = request.POST.get(
                        f"items-{sales_order_item_id}-notes", "")
                    file_attachment = request.FILES.get(
                        f"items-{sales_order_item_id}-file_attachment")
                    # pylint: disable=no-member
                    items_to_create.append({
                        "sales_order_item": sales_order_item,
                        "quantity_requested": quantity_requested,
                        "price": price,
                        "notes": notes,
                        "file_attachment": file_attachment,
                        "supplier": Suppliers.objects.get(
                            id=supplier_id) if supplier_id else None
                    })
                    print(
                        f"[Debug] Ítem válido para creación: "
                        f"{sales_order_item.description},"
                        f"Cantidad: {quantity_requested}, Precio: {price}")

        # Verificar si hay ítems válidos para crear
        print("[Debug] Ítems válidos para crear:", items_to_create)
        if not items_to_create:
            print("[Error] No se pueden crear órdenes sin ítems válidos")
            return JsonResponse(
                {"message": "Error: No se puede crear una orden de requerimiento sin ítems válidos.", "type": "error"},
                status=400)

        # Si alguna validación falló (como cantidad disponible <= 0
        # o cantidad solicitada > remaining_requirement),
        # no permitimos crear la orden
        if order_has_errors:
            return JsonResponse(
                {"message": "Error: No se puede crear la orden debido a los errores en los ítems.", "type": "error"},
                status=400)

        # Crear la orden de requerimiento
        requirement_order = RequirementOrder(
            sales_order=sales_order,
            user=request.user,
            requested_date=request.POST.get("requested_date"),
            notes=request.POST.get("notes")
        )
        requirement_order.save()
        print("[Debug] Orden de requerimiento creada.")

        # Crear cada ítem en la orden de requerimiento
        for item_data in items_to_create:
            item = RequirementOrderItem(
                requirement_order=requirement_order,
                sales_order_item=item_data["sales_order_item"],
                quantity_requested=item_data["quantity_requested"],
                price=item_data["price"],
                notes=item_data["notes"],
                file_attachment=item_data["file_attachment"],
                supplier=item_data["supplier"]
            )
            item.save()
            print(
                f"[Debug] Ítem {item.sales_order_item.description}"
                f"guardado con cantidad {item.quantity_requested}.")

        return JsonResponse(
            {
                "message": "Orden de Requerimiento creada exitosamente.",
                "type": "success"}
            )
    # pylint: disable=broad-except
    except Exception as e:
        print(
            f"[Error] Excepción al crear "
            f"la orden de requerimiento: {str(e)}")
        return JsonResponse(
            {
                "message": f"Error inesperado: {str(e)}",
                "type": "error"},
            status=400)


def export_requirement_order(order_id):
    """Export a requirement order to an Excel"""
    # Obtener la orden de requerimiento que quieres exportar
    order = get_object_or_404(
        RequirementOrder,
        id=order_id
    )

    # Crear un libro de trabajo (workbook)
    wb = Workbook()
    ws = wb.active
    ws.title = "Orden de Requerimiento"

    # Colores para la información general
    general_info_fill = PatternFill(
        start_color="89b5fa",
        end_color="89b5fa",
        fill_type="solid")
    general_info_alignment = Alignment(
        horizontal="left",
        vertical="center")  # Alineación izquierda

    # Agregar información general de la orden de requerimiento
    general_info = [
        ("Orden de Venta:", str(order.sales_order)),
        ("Fecha Solicitada:", order.requested_date),
        (
            "Fecha de Creación:",
            order.created_at.strftime("%d/%m/%Y")),
        (
            "Detalle del Requerimiento:",
            order.notes if order.notes else "No disponible"),
        (
            "Estado de la Orden:",
            dict(RequirementOrder.STATE_CHOICES).get(
                order.state, 'Desconocido')),
        ("Total de la Orden:", order.total_order),
    ]

    # Escribir la información general en las primeras filas con color y formato
    row_num = 1
    for label, value in general_info:
        ws.append([label, value])
        # Obtener la fila actual
        row_cells = ws[row_num]
        for cell in row_cells:
            cell.fill = general_info_fill  # Aplicar el color de fondo
            cell.alignment = general_info_alignment  # Alineación
        row_num += 1

    # Agregar una fila vacía para separar la información
    # general de los detalles de los artículos
    ws.append([])

    # Agregar encabezados para los detalles de los artículos
    headers = [
        "Item",
        "Info",
        "Cantidad",
        "Precio Unitario",
        "Precio Total",
        "Estado"
    ]
    ws.append(headers)

    # Aplicar formato a los encabezados (colorear y centrar el texto)
    header_fill = PatternFill(
        start_color="4CAF50",
        end_color="4CAF50",
        fill_type="solid")  # Verde
    header_font = Font(
        color="FFFFFF",
        bold=True)  # Texto blanco y en negrita
    header_alignment = Alignment(
        horizontal="center",
        vertical="center"
        )  # Centrado

    # Aplicar formato a todas las celdas de los encabezados (primera fila)
    for cell in ws[ws.max_row]:  # Primera fila (encabezados)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Definir los colores para los estados
    estado_colors = {
        'L': PatternFill(
            start_color="90EE90",
            end_color="90EE90",
            fill_type="solid"),  # Verde claro - Listo
        'P': PatternFill(
            start_color="FFFF00",
            end_color="FFFF00",
            fill_type="solid"),  # Amarillo - Pendiente
        'C': PatternFill(
            start_color="FFA500",
            end_color="FFA500",
            fill_type="solid"),  # Naranja - Comprando
        'R': PatternFill(
            start_color="FF0000",
            end_color="FF0000",
            fill_type="solid"),  # Rojo - Rechazado
        'E': PatternFill(
            start_color="ADD8E6",
            end_color="ADD8E6",
            fill_type="solid"),  # Azul claro - Enviado
        'A': PatternFill(
            start_color="FFFFE0",
            end_color="FFFFE0",
            fill_type="solid"),  # Amarillo claro - Aceptado
    }

    # Llenar el archivo con los datos de la tabla
    for item in order.items.all():
        # Se calcula el total_price como cantidad * precio
        total_price = item.quantity_requested * item.price if item.price else 0
        estado = dict(
            RequirementOrderItem.ESTADO_CHOICES).get(
                item.estado, 'Desconocido')

        # Crear una fila para agregar
        row = [
            item.sales_order_item.description,
            item.notes if item.notes else "",
            item.quantity_requested,
            item.price if item.price else "",
            total_price,
            estado
        ]

        # Agregar la fila a la hoja
        ws.append(row)

        # Obtener la fila recién agregada (última fila)
        row_cells = ws[ws.max_row]  # Accede a la última fila (nueva)

        # Aplicar el color a toda la fila
        if item.estado in estado_colors:
            for cell in row_cells:
                cell.fill = estado_colors[item.estado]

    # Ajustar el ancho de las columnas basado en el contenido
    # Desde la columna 1 hasta la columna de los encabezados
    for col in range(1, len(headers) + 1):
        max_length = 0
        # Obtenemos la columna actual
        column = ws.column_dimensions[get_column_letter(col)]
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    # Obtener el largo del valor más largo
                    max_length = max(
                        max_length,
                        len(str(cell.value)))
        column.width = max_length + 2  # Ajustar el ancho de la columna

    # Crear una respuesta HTTP con el archivo Excel como un adjunto
    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        )
    response['Content-Disposition'] = (
        f'attachment; filename=orden_requerimiento_{order_id}.xlsx'
        )

    # Guardar el libro de trabajo en la respuesta
    wb.save(response)

    return response
