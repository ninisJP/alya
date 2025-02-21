# See LICENSE file for copyright and license details.
"""
Budget utils
"""
import decimal
from decimal import Decimal

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

from django.conf import settings
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from .models import Budget, BudgetItem, CatalogItem


def export_budget_report_to_excel(request, pk):
    """
    Export budget report to excel
    """

    budget = get_object_or_404(Budget, pk=pk)

    # Load excel template
    file_path = f'{settings.BASE_DIR}/static/plantilla.xlsx'
    plantilla = openpyxl.load_workbook(file_path)

    items_by_category = {}
    for item in budget.items.all():
        categoria = item.item.category
        if categoria not in items_by_category:
            items_by_category[categoria] = []
        items_by_category[categoria].append(item)

    # Sheet by category
    for categoria, items in items_by_category.items():
        ws_categoria = plantilla.create_sheet(title=f"{categoria}")
        _crear_hoja_presupuesto(
            ws_categoria,
            budget,
            {categoria: items},
            simbolo='S/'
        )

    ws_resumen = plantilla.create_sheet(title="Resumen Final")
    _crear_hoja_resumen(ws_resumen, budget, items_by_category)
    _create_quote(plantilla, budget)
    _create_budget(plantilla, budget, items_by_category)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename="cotizacion_{budget.budget_name}.xlsx"'
    )
    plantilla.save(response)

    return response


def _crear_hoja_presupuesto(ws, budget, items_by_category, simbolo='S/'):
    """
    Create budget sheet in excel
    """

    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    cell_font = Font(name="Calibri", size=11)
    alignment_center = Alignment(horizontal="center", vertical="center")
    fill_blue = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Título del presupuesto
    ws['A1'] = f'PRESUPUESTO: {budget.budget_name} ({simbolo})'
    ws['A1'].font = Font(bold=True, size=16, name="Calibri")
    ws['A1'].alignment = alignment_center

    # Encabezados de la tabla
    headers = [
        'ITEM',
        'DESCRIPCION DE ITEM',
        'CATEGORÍA',
        'CANTIDAD',
        'MONEDA',
        'PRECIO UNITARIO',
        'SUBTOTAL'
    ]
    ws.append(headers)
    for cell in ws[2]:  # 2nd row
        cell.font = header_font
        cell.fill = fill_blue
        cell.alignment = alignment_center
        cell.border = thin_border

    # Items data
    for items in items_by_category.values():
        for item in items:
            subtotal = item.total_price
            row = [
                item.item.sap,
                item.item.description,
                item.item.category,
                item.quantity,
                item.coin,
                item.custom_price,
                subtotal
            ]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.font = cell_font
                cell.alignment = alignment_center
                cell.border = thin_border

    ws.append([''])
    resumen = []

    # Zero division
    if budget.budget_expenses != 0:
        percent = int(budget.budget_expenses) / 100
        print(f' heli {percent}')
        utilidad = budget.budget_price * Decimal(percent)
        total_final = budget.budget_price - utilidad
    else:
        utilidad = 0
        total_final = budget.budget_price

    resumen = [
        ('TOTAL PARCIAL', f'S/ {budget.budget_price:.2f}'),
        ('GASTOS ADMINISTRATIVOS', f'%{budget.budget_expenses}'),
        ('UTILIDAD', f'S/ {utilidad:.2f}'),
        ('TOTAL FINAL', f'S/ {total_final:.2f}')
    ]

    for label, value in resumen:
        ws.append([label, '', '', '', '', value])
        for cell in ws[ws.max_row]:
            cell.font = cell_font if label != 'TOTAL FINAL' else Font(
                bold=True,
                size=12,
                color="FF0000"
            )
            cell.alignment = alignment_center
            cell.border = thin_border
            if label == 'TOTAL FINAL':
                cell.fill = PatternFill(
                    start_color="FFFF00",
                    end_color="FFFF00",
                    fill_type="solid"
                )

    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        adjusted_width = max_length + 2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width


def _crear_hoja_resumen(ws, budget, items_by_category):
    """
    Make resumen sheet
    """

    header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    cell_font = Font(name="Calibri", size=11)
    alignment_center = Alignment(horizontal="center", vertical="center")
    fill_blue = PatternFill(
        start_color="4F81BD",
        end_color="4F81BD",
        fill_type="solid"
    )
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws['A1'] = f'RESUMEN FINAL DEL PRESUPUESTO: {budget.budget_name}'
    ws['A1'].font = Font(bold=True, size=16, name="Calibri")
    ws['A1'].alignment = alignment_center

    headers = [
        'ITEM',
        'DESCRIPCIÓN',
        'CATEGORÍA',
        'CANTIDAD',
        'MONEDA',
        'PRECIO UNITARIO',
        'SUBTOTAL'
    ]

    ws.append(headers)
    for cell in ws[2]:
        cell.font = header_font
        cell.fill = fill_blue
        cell.alignment = alignment_center
        cell.border = thin_border

    for categoria, items in items_by_category.items():
        ws.append([categoria])
        for item in items:
            subtotal = item.total_price
            row = [
                item.item.sap,
                item.item.description,
                item.item.category,
                item.quantity,
                item.coin,
                item.custom_price,
                subtotal,
            ]
            ws.append(row)
            for cell in ws[ws.max_row]:
                cell.font = cell_font
                cell.alignment = alignment_center
                cell.border = thin_border

    ws.append([''])
    resumen = []

    # Zero divition
    if budget.budget_expenses != 0:
        percent = int(budget.budget_expenses) / 100
        print(f' heli {percent}')
        utilidad = budget.budget_price * Decimal(percent)
        total_final = budget.budget_price - utilidad
    else:
        utilidad = 0
        total_final = budget.budget_price

    resumen = [
        ('TOTAL PARCIAL', f'S/ {budget.budget_price:.2f}'),
        ('GASTOS ADMINISTRATIVOS', f'%{budget.budget_expenses}'),
        ('UTILIDAD', f'S/ {utilidad:.2f}'),
        ('TOTAL FINAL', f'S/ {total_final:.2f}')
    ]

    for label, value in resumen:
        ws.append([label, '', '', '', '', value])
        for cell in ws[ws.max_row]:
            cell.font = cell_font if label != 'TOTAL FINAL' else Font(
                bold=True,
                size=12,
                color="FF0000"
            )
            cell.alignment = alignment_center
            cell.border = thin_border
            if label == 'TOTAL FINAL':
                cell.fill = PatternFill(
                    start_color="FFFF00",
                    end_color="FFFF00",
                    fill_type="solid"
                )

    # Column size
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        adjusted_width = max_length + 2
        ws.column_dimensions[column[0].column_letter].width = adjusted_width


def _create_budget(plantilla, budget, items_by_category):
    """
    Create budget in excel.
    """

    sheet = plantilla['PRESUPUESTO']

    title_cell = sheet['B1']
    title_cell.value = f'PRESUPUESTO: {budget.budget_name}'
    title_cell.font = Font(name="Calibri", size=6, bold=True)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = [
        'ITEM',
        'DESCRIPCION DE PARTIDA',
        'UNIDAD',
        'CANTIDAD',
        'MONEDA',
        'P.U.',
        'SUB TOTAL',
        'TOTAL'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col_num)
        cell.value = header
        cell.font = Font(name="Calibri", size=6, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(
            start_color="4F81BD",
            end_color="4F81BD",
            fill_type="solid"
        )
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    current_row = 3
    for categoria, items in items_by_category.items():
        category_cell = sheet.cell(
            row=current_row,
            column=1,
            value=f'{categoria}'
        )
        category_cell.font = Font(name="Calibri", size=6, bold=True)
        category_cell.alignment = Alignment(
            horizontal="left",
            vertical="center"
        )
        current_row += 1

        for item in items:
            subtotal = item.total_price
            row = [
                item.item.sap,
                item.item.description,
                item.item.unit,
                item.quantity,
                item.coin,
                item.custom_price,
                subtotal,
                subtotal,
            ]

            for col_num, value in enumerate(row, 1):
                cell = sheet.cell(row=current_row, column=col_num, value=value)
                cell.font = Font(name="Calibri", size=6)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            current_row += 1

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                # TODO: check except pass
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                # TODO: check except pass
                pass
        adjusted_width = max_length + 2
        sheet.column_dimensions[column].width = adjusted_width

    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value)
        adjusted_width = max_length + 2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width


def _create_quote(plantilla, budget):
    """
    Create quote in excel
    """

    sheet = plantilla['COTIZACION']
    sheet['C18'] = f'{budget.client}'
    sheet['C19'] = f'{budget.client.primary_contact}'
    sheet['C20'] = f'{budget.client.email}'
    sheet['C21'] = f'{budget.client.phone}'

    sheet['G20'] = f'{budget.budget_number}'
    sheet['G21'] = f'{budget.budget_date}'

    sheet['C27'] = f'{budget.budget_name}'
    sheet['G30'] = f'{budget.budget_final_price}'

    sheet['D35'] = f'{budget.budget_deliverytime}'
    sheet['D36'] = f'{budget.budget_servicetime}'
    sheet['D38'] = f'{budget.budget_warrantytime}'


def process_budget_excel(excel_file, budget_id):
    """
    Add budget process.
    """

    budget = Budget.objects.get(id=budget_id)

    xls = pd.ExcelFile(excel_file)
    if 'listado' not in xls.sheet_names:
        raise ValueError("El archivo no contiene una hoja llamada 'listado'.")

    df = pd.read_excel(xls, sheet_name='listado')

    expected_columns = ['DESCRIPCION DE PARTIDA', 'UND', 'CANT.', 'P.U.']
    if not all(col in df.columns for col in expected_columns):
        raise ValueError("El archivo Excel no tiene el formato esperado.")

    for index, row in df.iterrows():
        description = row['DESCRIPCION DE PARTIDA']
        unit = row['UND']
        quantity = row['CANT.']
        custom_price = row['P.U.']

        catalog_item = CatalogItem.objects.filter(
            description__iexact=description
        ).first()

        if catalog_item:
            BudgetItem.objects.create(
                budget=budget,
                item=catalog_item,
                quantity=quantity,
                custom_price=custom_price,
                custom_price_per_day=catalog_item.price_per_day,
                unit=unit
            )
        else:
            continue


def determine_category(sap_code):
    """
    Return the category by SAP code
    """

    if sap_code.startswith('HER'):
        return CatalogItem.Category.HERRAMIENTA
    if sap_code.startswith('CON'):
        return CatalogItem.Category.CONSUMIBLE
    if sap_code.startswith('MOB'):
        return CatalogItem.Category.MANODEOBRA
    if sap_code.startswith('MAT'):
        return CatalogItem.Category.MATERIAL
    if sap_code.startswith('EPPS'):
        return CatalogItem.Category.EPPS
    if sap_code.startswith('EQU'):
        return CatalogItem.Category.EQUIPO

    return CatalogItem.Category.EQUIPO


def safe_convert_to_decimal(value):
    """
    Fix simbols from excel
    """
    try:
        if isinstance(value, str):
            value = value.replace(",", "")
            value = value.replace("S/", "").strip()
            value = value.replace("US$", "").strip()
        return Decimal(value)
    except (ValueError, decimal.InvalidOperation):
        return Decimal(0)


def process_sap_excel(excel_file, budget):
    """
    Process excel to SAP.
    """

    xls = pd.ExcelFile(excel_file)
    if 'listado' not in xls.sheet_names:
        raise ValueError("El archivo no contiene una hoja llamada 'listado'.")

    df = pd.read_excel(xls, sheet_name='listado')

    expected_columns = [
        'Número de artículo',
        'Descripción del artículo',
        'Nombre de unidad de medida',
        'Cantidad',
        'Precio por unidad',
        'Total (ML)',
        'Moneda'
    ]
    if not all(col in df.columns for col in expected_columns):
        raise ValueError("El archivo Excel no tiene el formato esperado.")

    df['Cantidad'] = df['Cantidad'].apply(safe_convert_to_decimal)
    df['Precio por unidad'] = df['Precio por unidad'].apply(
        safe_convert_to_decimal
    )
    df['Total (ML)'] = df['Total (ML)'].apply(safe_convert_to_decimal)
    df['Categoría'] = df['Número de artículo'].apply(determine_category)
    df = df.sort_values(by='Categoría')

    excel_total_sum = Decimal('0.00')
    processed_total_sum = Decimal('0.00')
    sap_codes_in_excel = set()

    with transaction.atomic():
        for index, row in df.iterrows():
            sap_code = row['Número de artículo']

            # Check SAP
            if pd.isna(sap_code) or pd.isnull(sap_code):
                continue

            sap_code = str(sap_code).strip()
            sap_codes_in_excel.add(sap_code)

            description = row['Descripción del artículo']
            unit = row['Nombre de unidad de medida']
            quantity = row['Cantidad']
            coin = row['Moneda']
            custom_price = row['Precio por unidad']
            total_price = row['Total (ML)']
            excel_total_sum += total_price

            # Search catalog item
            try:
                catalog_item = CatalogItem.objects.get(sap=sap_code)
                catalog_item.description = description
                catalog_item.unit = unit
                catalog_item.price = custom_price
                catalog_item.price_per_day = custom_price / Decimal(30) if custom_price else Decimal(0)
                catalog_item.save()
            except CatalogItem.DoesNotExist:
                # New item
                category = determine_category(sap_code)
                catalog_item = CatalogItem.objects.create(
                    sap=sap_code,
                    description=description,
                    unit=unit,
                    price=custom_price,
                    price_per_day=custom_price / Decimal(30) if custom_price else Decimal(0),
                    category=category
                )

            budget_item, created = BudgetItem.objects.get_or_create(
                budget=budget,
                item=catalog_item,
                coin=coin,
            )

            budget_item.quantity = quantity
            budget_item.custom_price = custom_price
            budget_item.unit = unit
            budget_item.total_price = total_price
            budget_item.coin = coin
            budget_item.save()

            processed_total_sum += budget_item.total_price

        current_sap_codes = set(
            budget.items.values_list(
                'item__sap',
                flat=True
            )
        )
        sap_codes_to_delete = current_sap_codes - sap_codes_in_excel

        if sap_codes_to_delete:
            items_to_delete = budget.items.filter(
                item__sap__in=sap_codes_to_delete
            )
            items_to_delete.delete()

        budget.budget_price = processed_total_sum.quantize(Decimal('0.01'))
        budget.save()
