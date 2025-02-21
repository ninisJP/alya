import traceback
import pdfplumber
import pandas as pd
import logging
import re
from decimal import Decimal, InvalidOperation
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Q, Sum
from django.forms import modelformset_factory
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.utils.timezone import localdate
from django.views.decorators.http import require_POST
from django.views.generic import ListView, DetailView, FormView
from alya.utils import send_state_change_email
from logistic_requirements.forms import RequirementOrderForm, RequirementOrderItemFormSet
from logistic_requirements.models import RequirementOrder, RequirementOrderItem
from logistic_suppliers.models import Suppliers
from .forms import (
    PurchaseOrderForm,
    PurchaseOrderItemForm,
    SalesOrderForm,
    ItemSalesOrderExcelForm,
    ItemSalesOrderForm,
    BankForm,
    UploadBankStatementForm,
    CollectionOrdersForm,
    PurchaseOrderSearchForm,
)
from .models import (
    Bank,
    BankStatements,
    CollectionOrders,
    PurchaseOrder,
    PurchaseOrderItem,
    Rendition,
    SalesOrder,
    SalesOrderItem,
)
from .utils import extraer_datos_pdf, procesar_archivo_excel
from datetime import datetime, timedelta
from django.shortcuts import render
from django.db.models import Prefetch
import json
from django.views.decorators.csrf import csrf_exempt

def salesorder(request):
    salesorders = SalesOrder.objects.all().order_by("-id")
    context = {'form': SalesOrderForm(), 'salesorders': salesorders}
    return render(request, 'salesorder/sales_index.html', context)

def salesorder_search(request):
    query = request.GET.get('q', '')  # Obtener la consulta de búsqueda desde el request

    # Si hay una búsqueda, filtrar las órdenes de venta basadas en la consulta
    if query:
        salesorders = SalesOrder.objects.filter(
            Q(sapcode__icontains=query) | Q(project__name__icontains=query) | Q(detail__icontains=query)
            ).order_by('-id')
    else:
        salesorders = SalesOrder.objects.all().order_by('-id')  # Mostrar todas las órdenes de venta si no hay búsqueda

    context = {'salesorders': salesorders , 'form': SalesOrderForm()}
    # Devolver solo el fragmento de la lista
    return render(request, 'salesorder/salesorder-list.html', context)

def create_salesorder(request):
    if request.method == 'POST':
        form = SalesOrderForm(request.POST)
        if form.is_valid():
            form.save()
            salesorders = SalesOrder.objects.all().order_by("-id")
            context = {'salesorders': salesorders}
            return render(request, 'salesorder/salesorder-list.html', context)
    else:
        form = SalesOrderForm()
    return render(request, 'salesorder/salesorder-form.html', {'form': form})

def edit_salesorder(request, salesorder_id):
    salesorder = get_object_or_404(SalesOrder, id=salesorder_id)
    if request.method == 'GET':
        form = SalesOrderForm(instance=salesorder)
        print(form)
        return render(request, 'salesorder/salesorder-edit.html', {'form': form, 'salesorder': salesorder})
    elif request.method == 'POST':
        form = SalesOrderForm(request.POST, instance=salesorder)
        if form.is_valid():
            form.save()
            salesorders = SalesOrder.objects.all().order_by("-id")
            return render(request, 'salesorder/salesorder-list.html', {'salesorders': salesorders})
    return HttpResponse(status=405)

def delete_salesorder(request, salesorder_id):
    salesorder = get_object_or_404(SalesOrder, id=salesorder_id)
    if request.method == 'DELETE':
        salesorder.delete()
        salesorders = SalesOrder.objects.all().order_by("-id")
        return render(request, 'salesorder/salesorder-list.html', {'salesorders': salesorders})
    return HttpResponse(status=405)

def items_salesorder(request, salesorder_id):
    salesorder = get_object_or_404(SalesOrder, id=salesorder_id)
    items = SalesOrderItem.objects.filter(salesorder=salesorder)

    if request.method == "POST":
        form = ItemSalesOrderForm(request.POST)
        excel_form = ItemSalesOrderExcelForm(request.POST, request.FILES)

        if form.is_valid():
            item = form.save(commit=False)
            item.salesorder = salesorder
            item.save()
            items = SalesOrderItem.objects.filter(salesorder=salesorder)
            return render(request, 'itemsalesorder/item-salesorder-list.html', {'items': items})

        if excel_form.is_valid():
            archivo_excel = request.FILES['archivo_excel']
            procesar_archivo_excel(archivo_excel, salesorder_id)
            items = SalesOrderItem.objects.filter(salesorder=salesorder)
            return render(request, 'itemsalesorder/item-salesorder-list.html', {'items': items})

    else:
        form = ItemSalesOrderForm()
        excel_form = ItemSalesOrderExcelForm()

    context = {
        'salesorder': salesorder,
        'items': items,
        'form': form,
        'excel_form': excel_form
    }

    return render(request, 'itemsalesorder/item-salesorder.html', context)

def quick_create_purchaseorder(request, salesorder_id):
    salesorder = get_object_or_404(SalesOrder, id=salesorder_id)
    items = SalesOrderItem.objects.filter(salesorder=salesorder)

    if request.method == "POST":
        # Obtenemos el ID del item desde el formulario y la cantidad
        item_id = request.POST.get("item_id")
        quantity_requested = int(request.POST.get("quantity_requested"))

        sales_order_item = get_object_or_404(SalesOrderItem, id=item_id)

        # Creamos la PurchaseOrder basada en los detalles del SalesOrder
        purchase_order = PurchaseOrder.objects.create(
            salesorder=salesorder,
            description=f"Orden de Compra rápida para {sales_order_item.description}",
            requested_date=request.POST.get('requested_date', None),  # Opcional
            scheduled_date=request.POST.get('scheduled_date', None),  # Opcional
            requested_by=request.user.username  # Asumimos que el usuario actual está creando la orden
        )

        # Creamos el PurchaseOrderItem basado en el SalesOrderItem
        purchase_order_item = PurchaseOrderItem.objects.create(
            purchaseorder=purchase_order,
            sales_order_item=sales_order_item,
            sap_code=sales_order_item.sap_code,
            quantity_requested=quantity_requested,  # Usamos la cantidad ingresada por el usuario
            price=sales_order_item.price,
            price_total=sales_order_item.price * quantity_requested,
            supplier=None  # Este campo puede ser opcional o predeterminado
        )

        return redirect('purchaseorder_list')  # Redirige a la lista de órdenes de compra

    return render(request, 'itemsalesorder/item-salesorder.html', {
        'salesorder': salesorder,
        'items': items,
    })

def general_purchaseorder(request):
    # Obtener los parámetros del rango de fechas (si existen)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Si no se han proporcionado fechas, mostrar las órdenes sin scheduled_date
    if not start_date and not end_date:
        purchase_orders = PurchaseOrder.objects.filter(scheduled_date__isnull=True)
    else:
        # Si se proporcionan fechas, buscar entre esas dos fechas
        if not end_date:
            end_date = localdate()  # Si solo se proporciona la fecha de inicio, se asume que el rango termina hoy

        purchase_orders = PurchaseOrder.objects.filter(
            scheduled_date__range=[start_date, end_date]
        )

    context = {
        'purchase_orders': purchase_orders,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'purchaseorder/general_purchaseorder.html', context)

def purchaseorder_search(request):
    query = request.GET.get('q', '')

    if query:
        purchase_orders = PurchaseOrder.objects.filter(
            Q(salesorder__project__name__icontains=query)| Q(description__icontains=query)).order_by('-id')
    else:
        purchase_orders = PurchaseOrder.objects.all().order_by('-id')

    context = {'purchase_orders': purchase_orders , 'form': PurchaseOrderSearchForm()}
    
    return render(request, 'purchaseorder/purchaseorder_list_partial.html', context)

# Ordenes de compra
def purchase_orders(request, salesorder_id):
    salesorder = get_object_or_404(SalesOrder, id=salesorder_id)
    purchase_orders = PurchaseOrder.objects.filter(salesorder=salesorder)

    context = {
        'salesorder': salesorder,
        'purchase_orders': purchase_orders,
    }

    return render(request, 'purchaseorder/purchaseorder_list.html', context)

def purchase_orders_detail(request, purchaseorder_id):
    salesorder = get_object_or_404(SalesOrder, id=purchaseorder_id)
    items = salesorder.items.all()
    
    for item in items:
        item.diff = item.amount - item.remaining_requirement
    context = {
        'salesorder': salesorder,
        'items': items,
    }
    return render(request, 'purchaseorder/purchase_order_detail.html', context)

def edit_purchase_order(request, order_id):
    order = get_object_or_404(PurchaseOrder, id=order_id)
    ItemFormSet = modelformset_factory(PurchaseOrderItem, form=PurchaseOrderItemForm, extra=0)

    if request.method == 'POST':
        order_form = PurchaseOrderForm(request.POST, instance=order)
        item_formset = ItemFormSet(request.POST, queryset=PurchaseOrderItem.objects.filter(purchaseorder=order))

        if order_form.is_valid() and item_formset.is_valid():
            order_form.save()

            # Guardar cada ítem y recalcular el price_total
            items = item_formset.save(commit=False)
            for item in items:
                item.price_total = item.price * item.quantity_requested  # Recalcular
                item.save()

            return render(request, 'purchaseorder/purchaseorder_partial.html', {'order': order})
    else:
        order_form = PurchaseOrderForm(instance=order)
        item_formset = ItemFormSet(queryset=PurchaseOrderItem.objects.filter(purchaseorder=order))

    return render(request, 'purchaseorder/purchaseorder_form_partial.html', {
        'order_form': order_form,
        'item_formset': item_formset,
        'order': order,
    })

def delete_purchase_order(request, order_id):
    # Obtener la orden de compra a eliminar
    order = get_object_or_404(PurchaseOrder, id=order_id)

    # Recuperar los ítems de la orden de compra que están en estado 'C' y han sido procesados
    items_comprando = RequirementOrderItem.objects.filter(
        requirement_order__purchase_order=order,  # Aquí usamos 'requirement_order__purchase_order'
        estado='C',
        purchase_order_created=True
    )

    # Revertir el estado de 'purchase_order_created' a False y cambiar el estado a 'P' para los ítems
    for item in items_comprando:
        item.purchase_order_created = False  # Revertir la creación de la orden de compra
        item.estado = 'P'  # Cambiar el estado a 'Pendiente'
        item.save()

    # Eliminar la orden de compra
    order.delete()

    # Retornar un mensaje simple en HTML
    return HttpResponse('<div class="alert alert-success">Orden de compra eliminada con éxito. Los ítems han sido revertidos a estado Pendiente.</div>', content_type="text/html")

# Bank
def index_bank(request):
    if request.method == 'POST':
        form = BankForm(request.POST)
        if form.is_valid():
            print("Formulario válido. Banco:", form.cleaned_data.get('bank_name'))
            form.save()

            if request.headers.get('HX-Request'):  # Si es una petición HTMX
                context = {'banks': Bank.objects.all().order_by('-id')}
                return render(request, 'bank/bank_list.html', context)

            context = {'form': BankForm(), 'banks': Bank.objects.all()}
            return render(request, 'bank/bank_index.html', context)

        if request.headers.get('HX-Request'):  # Manejar los errores en HTMX
            return HttpResponse("Error en el formulario", status=400)

        return render(request, 'client/partials/failure_client.html')

    # En caso de GET request, mostrar el formulario vacío y la lista de bancos
    form = BankForm()
    context = {
        'form': form,
        'banks': Bank.objects.all().order_by('-id'),
        'pagina_activa': 'bancos'
    }
    return render(request, 'bank/bank_index.html', context)

def edit_bank(request, bank_id):
    bank = get_object_or_404(Bank, id=bank_id)

    if request.method == 'POST':
        form = BankForm(request.POST, instance=bank)
        if form.is_valid():
            bank = form.save()

            if request.headers.get('HX-Request'):
                banks = Bank.objects.all().order_by('-id')
                return render(request, 'bank/bank_list.html', {'bank': bank, 'banks': banks})

            return redirect('bank_index')

        if request.headers.get('HX-Request'):  # Manejar los errores en HTMX
            return HttpResponse("Error en el formulario", status=400)

        return render(request, 'client/partials/failure_client.html')

    # Corregir aquí asegurándonos de pasar el objeto `bank` al contexto
    form = BankForm(instance=bank)
    banks = Bank.objects.all().order_by('-id')
    context = {
        'form': form,
        'banks': banks,
        'bank': bank
    }
    return render(request, 'bank/bank_edit.html', context)

# Bank Delete
def delete_bank(request, bank_id):
    bank = get_object_or_404(Bank, id=bank_id)

    if request.method == 'POST':
        bank.delete()

        if request.headers.get('HX-Request'):  # Si es una petición HTMX
            return HttpResponse(status=204)

        return redirect('bank_index')  # Redirigir después de eliminar

    return render(request, 'bank/bank_delete.html', {'bank': bank})

# BankStatements
def bank_statements(request, bank_id):
    bank = get_object_or_404(Bank, pk=bank_id)
    statements = BankStatements.objects.filter(bank=bank)
    return render(request, 'bankstatement/BankStatements_list.html', {'bank': bank, 'statements': statements})

class BankStatementUploadView(FormView):
    template_name = 'bankstatement/upload_bank_statements.html'
    form_class = UploadBankStatementForm
    success_url = reverse_lazy('bank_index')

    def form_valid(self, form):
        bank = form.cleaned_data['bank']
        bank_name = bank.bank_name.strip()
        uploaded_file = self.request.FILES['excel_file']

        try:
            # Leer todas las hojas del archivo Excel
            sheets = pd.read_excel(uploaded_file, sheet_name=None, engine='openpyxl', header=2)
            print("Nombres de hojas en el archivo:", sheets.keys())

            # Verificar si el nombre del banco coincide con alguna hoja
            if bank_name in sheets:
                df = sheets[bank_name]
                print("Columnas en la hoja antes de ajustar:", df.columns)

                # Renombrar columnas para que coincidan con los campos del modelo
                column_mapping = {
                    'F.Operac.': 'operation_date',
                    'Referencia': 'reference',
                    'Importe': 'amount',
                    'ITF': 'itf',
                    'Num.Mvto': 'number_moviment',
                }
                df.rename(columns=column_mapping, inplace=True)

                # Filtrar columnas necesarias
                df = df[['operation_date', 'reference', 'amount', 'itf', 'number_moviment']]
                print("Columnas en la hoja después de ajustar:", df.columns)

                # Eliminar filas con valores vacíos en las columnas clave
                initial_count = len(df)
                df.dropna(subset=['operation_date', 'reference', 'amount', 'number_moviment'], inplace=True)
                print(f"Total de filas antes de filtrar vacíos: {initial_count}, después: {len(df)}")

                # Procesar cada fila
                for _, row in df.iterrows():
                    try:
                        # Convertir la fecha utilizando pd.to_datetime()
                        operation_date = pd.to_datetime(row['operation_date']).date()
                    except Exception as e:
                        messages.warning(self.request, f"Fecha inválida en número de movimiento {row['number_moviment']}: {row['operation_date']}")
                        continue  # Saltar a la siguiente fila en caso de error

                    # Evitar duplicados
                    if not BankStatements.objects.filter(
                        bank=bank,
                        operation_date=operation_date,
                        number_moviment=row['number_moviment']
                    ).exists():
                        # Crear registro
                        BankStatements.objects.create(
                            bank=bank,
                            operation_date=operation_date,
                            reference=row['reference'],
                            amount=row['amount'],
                            itf=row['itf'],
                            number_moviment=row['number_moviment']
                        )
                    else:
                        print(f"Registro duplicado encontrado para número de movimiento: {row['number_moviment']}")

                # Mensaje de éxito
                messages.success(self.request, f'Extractos bancarios para {bank_name} subidos exitosamente.')
            else:
                messages.error(self.request, f'No se encontró una hoja para el banco: {bank_name}.')

        except Exception as e:
            print("Error procesando archivo:", e)
            messages.error(self.request, f'Hubo un error procesando el archivo: {e}')

        return super().form_valid(form)

@require_POST
def assign_bank_statement(request, item_id, statement_id):
    try:
        # Obtener el ítem y el extracto bancario
        item = PurchaseOrderItem.objects.get(id=item_id)
        bank_statement = BankStatements.objects.get(id=statement_id)

        # Asignar el extracto bancario al ítem
        item.bank_statement = bank_statement
        item.mov_number = bank_statement.number_moviment
        item.save()

        return JsonResponse({'success': True, 'message': 'Conciliación realizada correctamente.'})
    except PurchaseOrderItem.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Ítem no encontrado.'})
    except BankStatements.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Extracto bancario no encontrado.'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'})

# Bank loands and credit cards
def bank_loans(request):
    pass
    # prestamo = BankLoan.objects.filter(credit_type = 'prestamo')
    # credito = BankLoan.objects.filter(credit_type = 'credito')

    # context = {
    #     'prestamos': prestamos,
    #     'creditos': creditos
    # }

    # return render(request, 'bank_loans.html', )

#caja chica
def petty_cash(request):
    # Obtener la fecha de hoy según la zona horaria configurada
    today = localdate()

    # Obtenemos los parámetros del rango de fechas (si existen)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Si no se han proporcionado fechas, mostrar ítems del día actual
    if not start_date and not end_date:
        items = PurchaseOrderItem.objects.filter(purchaseorder__scheduled_date=today).select_related(
            'purchaseorder', 'sales_order_item__salesorder', 'supplier'
        )
    else:
        # Si se proporcionan fechas, buscar entre esas dos fechas
        if not end_date:
            end_date = today  # Si solo hay fecha de inicio, el rango termina en el día actual

        if not start_date:
            start_date = today

        items = PurchaseOrderItem.objects.filter(
            purchase_date__range=[start_date, end_date]
        ).select_related('purchaseorder', 'sales_order_item__salesorder', 'supplier')

    context = {
        'items': items,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'pettycash/petty_cash_items.html', context)

def petty_cash_state(request):
    # Obtener los parámetros de filtro
    payment_status = request.GET.get('status')  # Puede ser "Pagado" o "No Pagado"
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Inicializar el queryset base
    items = PurchaseOrderItem.objects.select_related(
        'purchaseorder', 'sales_order_item__salesorder', 'supplier',
    )

    # Filtrar por estado de pago si está definido
    if payment_status:
        items = items.filter(payment_status=payment_status)

    # Si no se filtra por estado de pago, aplicar el filtro de fechas
    elif start_date or end_date:
        if not end_date:
            end_date = localdate()  # Fecha actual como final si no se especifica
        if not start_date:
            start_date = localdate()  # Fecha actual como inicio si no se especifica

        items = items.filter(purchase_date__range=[start_date, end_date])

    context = {
        'items': items,
        'start_date': start_date,
        'end_date': end_date,
        'payment_status': payment_status,
        'class_pay_choices': PurchaseOrderItem.CLASS_PAY_CHOICES,
        'type_pay_choices': PurchaseOrderItem.TYPE_PAY_CHOICES,
    }
    return render(request, 'pettycash/petty_cash_state.html', context)

def update_payment_status(request, item_id):
    item = get_object_or_404(PurchaseOrderItem, id=item_id)
    item.payment_status = 'Pagado' if item.payment_status == 'No Pagado' else 'No Pagado'
    item.save()
    return JsonResponse({'status': item.payment_status})

@csrf_exempt
def update_field(request, item_id):
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            field = data.get('field')
            value = data.get('value')

            item = get_object_or_404(PurchaseOrderItem, id=item_id)

            if field in ['class_pay', 'type_pay']:
                setattr(item, field, value)
                item.save()
                return JsonResponse({'success': True, 'field': field, 'value': value})

            return JsonResponse({'success': False, 'error': 'Campo no válido.'}, status=400)

        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# Import requirements views
class AccountingRequirementOrderListView(ListView):
    model = RequirementOrder
    template_name = 'requirements/requirement_order_list.html'
    context_object_name = 'requirement_orders'

    def get_queryset(self):
        # Obtiene el parámetro "filter" desde la URL
        filter_type = self.request.GET.get('filter', 'no_revisado')

        if filter_type == 'all':
            # Si el filtro es "all", retorna todas las órdenes
            return RequirementOrder.objects.all().order_by('-id').prefetch_related('items')
        else:
            # Por defecto muestra solo las órdenes "NO REVISADO"
            return RequirementOrder.objects.filter(state="NO REVISADO").order_by('-id').prefetch_related('items')

def accounting_requirement_order_detail_view(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    items = requirement_order.items.all()
    suppliers = Suppliers.objects.all()
    return render(request, 'requirements/requirement_order_detail.html', {
        'requirement_order': requirement_order,
        'items': items,
        'suppliers': suppliers,
    })

@require_POST
def update_requirement_order_items(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    updated_items = []

    new_state = request.POST.get('requirement_order_state')
    if new_state in dict(RequirementOrder.STATE_CHOICES):
        requirement_order.state = new_state

    # Recorrer los ítems de la orden y actualizar con los datos recibidos del request.POST
    for item in requirement_order.items.all():
        item.quantity_requested = request.POST.get(f'quantity_requested_{item.id}', item.quantity_requested)
        item.price = request.POST.get(f'price_{item.id}', item.price)
        item.notes = request.POST.get(f'notes_{item.id}', item.notes)
        item.supplier_id = request.POST.get(f'supplier_{item.id}')
        item.estado = request.POST.get(f'estado_{item.id}', item.estado)
        item.save()
        updated_items.append(item)

    requirement_order.save()

    return JsonResponse({'message': 'Elementos actualizados con éxito'}, status=200)

def logistic_create_purchase_order(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)

    # Verificar si ya existe una orden de compra para esta orden de requerimiento
    if requirement_order.purchase_order_created:
        error_message = "<div>Ya se se ha creado una Orden de Compra para esta Orden de Requerimiento.</div>"
        return HttpResponse(error_message, content_type="text/html")

    # Filtrar los ítems que están en estado "C"
    items_comprando = RequirementOrderItem.objects.filter(requirement_order=requirement_order, estado='C')

    if not items_comprando.exists():
        error_message = "<div>No hay ítems en estado 'Comprando' para crear una Orden de Compra.</div>"
        return HttpResponse(error_message, content_type="text/html")

    # Crear la PurchaseOrder
    with transaction.atomic():
        purchase_order = PurchaseOrder.objects.create(
            salesorder=requirement_order.sales_order,
            description=f"{requirement_order.notes} - {requirement_order.order_number}",
            requested_date=requirement_order.requested_date,
            requested_by=request.user.username if request.user else 'Desconocido',
            acepted=True
        )

        # Crear los PurchaseOrderItems asociados a los ítems comprando
        purchase_order_items = [
            PurchaseOrderItem(
                purchaseorder=purchase_order,
                sales_order_item=item.sales_order_item,
                sap_code=item.sap_code,
                quantity_requested=item.quantity_requested,
                price=item.price,
                price_total=item.total_price,
                notes=item.notes,
                supplier=item.supplier
            )
            for item in items_comprando
        ]
        PurchaseOrderItem.objects.bulk_create(purchase_order_items)

        # Actualizar la RequirementOrder para indicar que la orden de compra ha sido creada
        requirement_order.purchase_order_created = True
        requirement_order.save()

    # Respuesta de éxito en HTML
    success_message = f"<div>Orden de Compra creada para la Orden de Requerimiento #{requirement_order.order_number}.</div>"
    return HttpResponse(success_message, content_type="text/html")

#APROBAR REQUERIMIENTOS:
@require_POST
def update_requirement_order_state(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    new_state = request.POST.get('state')

    if new_state in dict(RequirementOrder.STATE_CHOICES):
        requirement_order.state = new_state
        requirement_order.save()
        # Mensaje de éxito
        message = 'La orden ha sido aprobada con éxito.' if new_state == 'APROBADO' else 'La orden ha sido rechazada con éxito.' #! ojito
        response_content = f"<div>{message}</div>"
        return HttpResponse(response_content, content_type="text/html")
    else:
        # Mensaje de error
        error_message = "<div>Estado inválido. Por favor, selecciona un estado válido.</div>"
        return HttpResponse(error_message, content_type="text/html")

def ajax_load_suppliers(request):
    term = request.GET.get('term', '')
    suppliers = Suppliers.objects.filter(name__icontains=term)[:20]
    supplier_list = [{'id': supplier.id, 'text': supplier.name} for supplier in suppliers]
    return JsonResponse({'results': supplier_list})

def purchase_conciliations(request):
    # Obtener los parámetros de fecha y nombre del banco
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    bank_name = request.GET.get('bank_name', '').strip()  # Tomar el nombre del banco o cadena vacía si no hay

    # Inicializar `items` y `bank_statements` como listas vacías para que la página cargue sin datos
    items = PurchaseOrderItem.objects.none()
    bank_statements = BankStatements.objects.none()

    # Solo aplicar filtros si se proporcionan `start_date`, `end_date`, o `bank_name`
    if start_date or end_date or bank_name:
        # Asignar valores por defecto si las fechas están vacías
        start_date = start_date or localdate()
        end_date = end_date or localdate()

        # Filtro para ítems sin número de movimiento y coincidencias en el nombre del banco
        items = PurchaseOrderItem.objects.filter(
                purchaseorder__scheduled_date__range=[start_date, end_date]
            ).filter(
                Q(mov_number__isnull=True) | Q(mov_number=''),  # Solo ítems sin número de movimiento
                supplier__bank__icontains=bank_name  # Coincidencia parcial en el nombre del banco del proveedor
            ).select_related('purchaseorder', 'sales_order_item__salesorder', 'supplier')

        # Filtro para los extractos bancarios por fecha y nombre del banco
        bank_statements = BankStatements.objects.filter(
            operation_date__range=[start_date, end_date],
            bank__bank_name__icontains=bank_name  # Coincidencia parcial en el nombre del banco de los extractos
        )

    context = {
        'items': items,
        'bank_statements': bank_statements,
        'start_date': start_date,
        'end_date': end_date,
        'bank_name': bank_name,
    }

    return render(request, 'conciliations/conciliations.html', context)

def report_conciliations(request):
    # Obtener el banco seleccionado desde el formulario
    selected_bank_id = request.GET.get('bank_id')
    selected_bank = None
    bank_statements = []

    # Obtener todos los bancos para el selector
    banks = Bank.objects.all()

    # Si se selecciona un banco, filtrar sus extractos
    if selected_bank_id:
        selected_bank = Bank.objects.filter(id=selected_bank_id).first()
        if selected_bank:
            # Obtener los extractos y las órdenes relacionadas
            bank_statements = BankStatements.objects.filter(bank=selected_bank).prefetch_related(
                Prefetch(
                    'conciliated_items',  # Relación definida en `PurchaseOrderItem`
                    queryset=PurchaseOrderItem.objects.select_related('purchaseorder', 'supplier'),
                )
            )

    context = {
        'banks': banks,
        'selected_bank': selected_bank,
        'bank_statements': bank_statements,
    }

    return render(request, 'conciliations/report_conciliations.html', context)

# renditions
def purchase_renditions(request):
    today = localdate()
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if not start_date and not end_date:
        items = PurchaseOrderItem.objects.filter(purchaseorder__scheduled_date=today).select_related(
            'purchaseorder', 'sales_order_item__salesorder', 'supplier'
        )
    else:
        if not end_date:
            end_date = today

        items = PurchaseOrderItem.objects.filter(
            purchaseorder__scheduled_date__range=[start_date, end_date]
        ).select_related('purchaseorder', 'sales_order_item__salesorder', 'supplier')

    # Calcular el total restante para cada item
    for item in items:
        total_renditions = item.renditions.aggregate(total=Sum('amount'))['total'] or 0
        item.total_remaining = item.price_total - total_renditions if item.price_total else 0

    context = {
        'items': items,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'renditions/renditions.html', context)

@require_POST
def add_rendition(request):
    # Extraer los datos del formulario
    item_id = request.POST.get('item_id')
    amount = request.POST.get('amount')
    series = request.POST.get('series')  # Obtener la serie del formulario
    correlative = request.POST.get('correlative')  # Obtener el correlativo del formulario
    date = request.POST.get('date')
    photo = request.FILES.get('photo')


    # Obtener el ítem o devolver un error 404 si no existe
    item = get_object_or_404(PurchaseOrderItem, id=item_id)

    try:
        # Crear la rendición con los datos recibidos
        rendition = Rendition(
            purchase_order_item=item,
            amount=amount,
            series=series,  # Guardar la serie
            correlative=correlative,  # Guardar el correlativo
            date=date if date else None,
            photo=photo,

        )

        # Intentar guardar la rendición
        rendition.save()

        # Si todo va bien, devolver un mensaje de éxito
        return JsonResponse({'status': 'success', 'message': 'Rendición añadida correctamente'})

    except ValidationError as e:
        # Captura el error y envía el primer mensaje de error de la lista
        return JsonResponse({'status': 'error', 'message': str(e.message or e.messages[0])})

# Listar órdenes de cobro y procesar carga de PDF
def collection_orders(request):
    if request.method == "POST":
        salesorder_id = request.POST.get("salesorder_id")
        pdf_file = request.FILES.get("pdf_file")

        if pdf_file and salesorder_id:
            orden_venta = get_object_or_404(SalesOrder, pk=salesorder_id)
            datos_pdf = extraer_datos_pdf(pdf_file)

            # Crear instancia de CollectionOrders con los datos extraídos
            collection_order = CollectionOrders(
                orden_venta=orden_venta,
                serie_correlativo=datos_pdf["serie_correlativo"],
                fecha_emision=datos_pdf["fecha_emision"],
                cliente=datos_pdf["cliente"],
                ruc_cliente=datos_pdf["ruc_cliente"],
                tipo_moneda=datos_pdf["tipo_moneda"],
                descripcion=datos_pdf["descripcion"],
                importe_total=datos_pdf["importe_total"],
                detraccion=datos_pdf["detraccion"],
                monto_neto_cobrar=datos_pdf["monto_neto_cobrar"],
                total_cuotas=datos_pdf["total_cuotas"],
                fecha_vencimiento=datos_pdf["fecha_vencimiento"],
            )
            collection_order.save()
            messages.success(request, "PDF procesado y datos guardados exitosamente.")

    # Obtener todas las SalesOrder para listarlas en la página
    sales_orders = SalesOrder.objects.all()
    context = {
        'sales_orders': sales_orders,
    }
    return render(request, 'collectionorders/collection_orders.html', context)

def collection_orders_search(request):
    query = request.GET.get('q','')
    if query:
        sales_orders = SalesOrder.objects.filter(Q(sapcode__icontains=query)).order_by('-id')
    else:
        sales_orders = SalesOrder.objects.all().order_by('-id')

    context = {'sales_orders':sales_orders,'form':SalesOrderForm()}
    return render(request,'collectionorders/collection_orders_list.html',context)

def collection_order_detail(request, salesorder_id):
    # Obtiene la SalesOrder y sus CollectionOrders asociadas
    orden_venta = get_object_or_404(SalesOrder, pk=salesorder_id)
    collection_orders = CollectionOrders.objects.filter(orden_venta=orden_venta)

    # Procesa el PDF si se carga uno
    if request.method == "POST":
        pdf_file = request.FILES.get("pdf_file")

        if pdf_file:
            datos_pdf = extraer_datos_pdf(pdf_file)

            # Crear una nueva instancia de CollectionOrders con los datos extraídos
            collection_order = CollectionOrders(
                orden_venta=orden_venta,
                serie_correlativo=datos_pdf["serie_correlativo"],
                fecha_emision=datos_pdf["fecha_emision"],
                cliente=datos_pdf["cliente"],
                ruc_cliente=datos_pdf["ruc_cliente"],
                tipo_moneda=datos_pdf["tipo_moneda"],
                descripcion=datos_pdf["descripcion"],
                importe_total=datos_pdf["importe_total"],
                detraccion=datos_pdf["detraccion"],
                monto_neto_cobrar=datos_pdf["monto_neto_cobrar"],
                total_cuotas=datos_pdf["total_cuotas"],
                fecha_vencimiento=datos_pdf["fecha_vencimiento"],
            )
            collection_order.save()
            messages.success(request, "PDF procesado y datos guardados exitosamente.")
            return redirect('collection_order_detail', salesorder_id=salesorder_id)

    context = {
        'orden_venta': orden_venta,
        'collection_orders': collection_orders,
    }
    return render(request, 'collectionorders/collection_order_detail.html', context)

# Eliminar una orden de cobro
def delete_collection_order(request, collection_order_id):
    collection_order = get_object_or_404(CollectionOrders, pk=collection_order_id)
    salesorder_id = collection_order.orden_venta.id
    collection_order.delete()
    return redirect(reverse('collection_orders', kwargs={'salesorder_id': salesorder_id}))

# Editar una orden de cobro
def edit_collection_order(request, collection_order_id):
    collection_order = get_object_or_404(CollectionOrders, pk=collection_order_id)

    if request.method == 'POST':
        form = CollectionOrdersForm(request.POST, instance=collection_order)
        if form.is_valid():
            form.save()
            return redirect(reverse('collection_orders', kwargs={'salesorder_id': collection_order.orden_venta.id}))
    else:
        form = CollectionOrdersForm(instance=collection_order)

    context = {
        'form': form,
        'collection_order': collection_order,
    }

    return render(request, 'collectionorders/edit_collection_order.html', context)

# Logistic Requirements Approved

class LogisticRequirementOrderListView(ListView):
    model = RequirementOrder
    template_name = 'logistic_requirements/logistic_index.html'
    context_object_name = 'requirement_orders'

    def get_queryset(self):
        # Obtener los parámetros de los filtros
        show_pending = self.request.GET.get('show_pending') == 'true'
        show_comprando = self.request.GET.get('show_comprando') == 'true'
        show_all = self.request.GET.get('show_all') == 'true'
        
        # Iniciar queryset con todas las órdenes APROBADAS por contabilidad
        queryset = RequirementOrder.objects.order_by('-id').prefetch_related('items') # filter(state='APROBADO')
        
        if show_all:
            queryset = queryset.distinct()
            print(show_all)
        elif show_comprando:
            # Mostrar solo las órdenes con ítems en estado Pendiente o Comprando
            queryset = queryset.filter(state='APROBADO').filter(
                    items__estado='C'# items__estado__in=['P', 'C']
            ).distinct()
            print(show_comprando)
        elif show_pending:
            # Mostrar solo las órdenes con ítems en estado Pendiente o Comprando
            queryset = queryset.filter(state='APROBADO').filter(
                    items__estado__in=['P', 'C']
            ).distinct()
            print(show_pending)
        else:
            queryset = queryset.filter(state='APROBADO').filter(
                items__estado='P' #items__estado__in=['P', 'C']
            ).distinct()
         
        
        # Calcular el estado general de cada orden
        for order in queryset:
            items = order.items.all()
            total_items = items.count()

            if total_items == 0:
                order.global_state = "No tiene ítems"
                continue

            ready_count = items.filter(estado='L').count()
            buying_count = items.filter(estado='C').count()
            pending_count = items.filter(estado='P').count()

            # Determinar el estado general
            if ready_count == total_items:
                order.global_state = "Listo"
            elif buying_count >= total_items / 2:
                order.global_state = "Comprando"
            elif pending_count > 0:
                order.global_state = "Pendiente"
            else:
                order.global_state = "Completado"  # Si no hay items pendientes, listos o comprando

        return queryset

from logistic_requirements.forms import RequirementOrderListForm
from logistic_inventory.models import Item

def logistic_search_requirement_order_list_view(request):
    query = request.GET.get('q', '')
    print(f"Search Query: {query}")  # Agregar log

    if query:
        requirement_orders = RequirementOrder.objects.filter(Q(order_number__icontains=query) | Q(estado__icontains=query)
                                                            | Q(notes__icontains=query)).order_by('-id')
    else:
        requirement_orders = RequirementOrder.objects.all().order_by('-id')
    
    print(requirement_orders)

    context = {'requirement_orders': requirement_orders, 'form': RequirementOrderListForm()}
    return render(request, 'logistic_list.html', context)

def logistic_requirement_order_detail_view(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    filtrar = request.GET.get('filtrar')

    # Filtrar los ítems en estado 'P' si se requiere, o cargar todos los ítems ordenados por categoría
    if filtrar == 'P':
        items = requirement_order.items.filter(estado='P').select_related('sales_order_item').order_by('sales_order_item__category')
    else:
        items = requirement_order.items.all().select_related('sales_order_item').order_by('sales_order_item__category')

    suppliers = Suppliers.objects.all()

    # Crear un diccionario de inventario disponible, usando sap como clave
    inventory_data = {item.item.sap: item.quantity for item in Item.objects.select_related('item').all()}

    # Agregar disponibilidad y tiempo de servicio a cada item en el queryset
    for item in items:
        item.disponible_inventario = inventory_data.get(item.sap_code, 0)

        # Calcular tiempo de servicio en horas solo si la categoría no es 'Material', 'Consumible', o 'Equipo'
        if item.sales_order_item.category not in ["Material", "Consumible", "Equipo"]:
            item.tiempo_servicio = requirement_order.sales_order.days * 8 * item.quantity_requested
        else:
            item.tiempo_servicio = None

    return render(request, 'logistic_requirements/logistic_detail.html', {
        'requirement_order': requirement_order,
        'items': items,
        'suppliers': suppliers,
        'filtrar': filtrar,
    })

@require_POST
def logistic_update_requirement_order_items(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    updated_items = []

    # Procesar cada ítem en la orden de requerimiento
    for item in requirement_order.items.all():
        item.quantity_requested = request.POST.get(f'quantity_requested_{item.id}', item.quantity_requested)
        item.price = request.POST.get(f'price_{item.id}', item.price)
        item.notes = request.POST.get(f'notes_{item.id}', item.notes)
        item.supplier_id = request.POST.get(f'supplier_{item.id}')
        item.estado = request.POST.get(f'estado_{item.id}', item.estado)

        # Guardar la fecha de la orden de compra
        date_str = request.POST.get(f'date_purchase_order_{item.id}')
        if date_str:
            try:
                item.date_purchase_order = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                item.date_purchase_order = None

        print(f"Item ID: {item.id}, Fecha recibida: {date_str}, Fecha procesada: {item.date_purchase_order}")
        updated_items.append(item)

    # Usar bulk_update para mejorar el rendimiento
    RequirementOrderItem.objects.bulk_update(
        updated_items, 
        ['quantity_requested', 'price', 'notes', 'supplier_id', 'estado', 'date_purchase_order']
    )

    # Recalcular remaining_requirement para todos los sales_order_items relacionados
    for item in updated_items:
        item.sales_order_item.update_remaining_requirement()
        print(f'Item ID: {item.id}, Fecha: {item.date_purchase_order}')


    # Retornar el mensaje directamente en HTML
    return HttpResponse(
        '<div class="alert alert-success" role="alert">Items actualizados con éxito</div>',
        status=200
    )

def logistic_create_purchase_order(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)

    # Verificar si ya existe una orden de compra para esta orden de requerimiento
    if requirement_order.purchase_order_created:
        # Si ya existe una orden de compra, la buscamos
        try:
            purchase_order = PurchaseOrder.objects.get(salesorder=requirement_order.sales_order, acepted=True)
        except PurchaseOrder.DoesNotExist:
            # Si no existe, creamos una nueva
            purchase_order = PurchaseOrder.objects.create(
                salesorder=requirement_order.sales_order,
                description=f"{requirement_order.notes} - {requirement_order.order_number}",
                requested_date=requirement_order.requested_date,
                requested_by=request.user.username if request.user else 'Desconocido',
                acepted=True
            )
            # Marcamos la RequirementOrder como procesada
            requirement_order.purchase_order_created = True
            requirement_order.save()
    else:
        # Si no existe, creamos una nueva
        purchase_order = PurchaseOrder.objects.create(
            salesorder=requirement_order.sales_order,
            description=f"{requirement_order.notes} - {requirement_order.order_number}",
            requested_date=requirement_order.requested_date,
            requested_by=request.user.username if request.user else 'Desconocido',
            acepted=True
        )
        # Marcamos la RequirementOrder como procesada
        requirement_order.purchase_order_created = True
        requirement_order.save()

    # Filtrar ítems en estado 'C' que no han sido procesados
    items_comprando = RequirementOrderItem.objects.filter(
        requirement_order=requirement_order, 
        estado='C',
        purchase_order_created=False  # Solo ítems no procesados
    )

    if not items_comprando.exists():
        error_message = "<div>No hay ítems en estado 'Comprando' sin orden de compra asociada.</div>"
        return HttpResponse(error_message, content_type="text/html")

    # Crear los PurchaseOrderItems asociados a los ítems comprando
    purchase_order_items = [
        PurchaseOrderItem(
            purchaseorder=purchase_order,
            sales_order_item=item.sales_order_item,
            sap_code=item.sap_code,
            quantity_requested=item.quantity_requested,
            price=item.price,
            price_total=item.total_price,
            notes=item.notes,
            supplier=item.supplier,
            purchase_date=item.date_purchase_order
        )
        for item in items_comprando
    ]
    PurchaseOrderItem.objects.bulk_create(purchase_order_items)

    # Actualizar el campo purchase_order_created a True para los ítems procesados
    for item in items_comprando:
        item.purchase_order_created = True
        item.save()

    # Respuesta de éxito en HTML
    success_message = f"<div>Ítems añadidos a la Orden de Compra para la Orden de Requerimiento #{requirement_order.order_number}.</div>"
    return HttpResponse(success_message, content_type="text/html")

@require_POST
def logistic_update_and_create_purchase_order(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    updated_items = []

    # Procesar cada ítem en la orden de requerimiento
    for item in requirement_order.items.all():
        item.quantity_requested = request.POST.get(f'quantity_requested_{item.id}', item.quantity_requested)
        item.price = request.POST.get(f'price_{item.id}', item.price)
        item.notes = request.POST.get(f'notes_{item.id}', item.notes)
        item.supplier_id = request.POST.get(f'supplier_{item.id}')
        item.estado = request.POST.get(f'estado_{item.id}', item.estado)

        # Guardar la fecha de la orden de compra
        date_str = request.POST.get(f'date_purchase_order_{item.id}')
        if date_str:
            try:
                item.date_purchase_order = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                item.date_purchase_order = None

        updated_items.append(item)

    # Usar bulk_update para mejorar el rendimiento
    RequirementOrderItem.objects.bulk_update(
        updated_items, 
        ['quantity_requested', 'price', 'notes', 'supplier_id', 'estado', 'date_purchase_order']
    )

    # Verificar si ya existe una PurchaseOrder asociada
    if not requirement_order.purchase_order:  # Si no tiene una orden de compra asociada
        purchase_order = PurchaseOrder.objects.create(
            salesorder=requirement_order.sales_order,  # Relacionado con la SalesOrder
            description=f"{requirement_order.notes} - {requirement_order.order_number}",
            requested_date=requirement_order.requested_date,
            requested_by=request.user.username if request.user else 'Desconocido',
            acepted=True
        )
        # Asociar la nueva PurchaseOrder con la RequirementOrder
        requirement_order.purchase_order = purchase_order
        requirement_order.purchase_order_created = True
        requirement_order.save()
    else:
        purchase_order = requirement_order.purchase_order  # Usar la PurchaseOrder existente

    # Filtrar ítems en estado 'C' que no han sido procesados
    items_comprando = RequirementOrderItem.objects.filter(
        requirement_order=requirement_order, 
        estado='C',
        purchase_order_created=False
    )

    if items_comprando.exists():
        purchase_order_items = [
            PurchaseOrderItem(
                purchaseorder=purchase_order,
                sales_order_item=item.sales_order_item,
                sap_code=item.sap_code,
                quantity_requested=item.quantity_requested,
                price=item.price,
                price_total=item.total_price,
                notes=item.notes,
                supplier=item.supplier,
                purchase_date=item.date_purchase_order
            )
            for item in items_comprando
        ]
        PurchaseOrderItem.objects.bulk_create(purchase_order_items)

        # Actualizar el campo purchase_order_created a True para los ítems procesados
        for item in items_comprando:
            item.purchase_order_created = True
            item.save()

    return HttpResponse(
        '<div class="alert alert-success" role="alert">Ítems actualizados y Orden de Compra creada con éxito</div>',
        status=200
    )

def ajax_load_suppliers(request):
    term = request.GET.get('term', '')
    suppliers = Suppliers.objects.filter(name__icontains=term)[:20]
    supplier_list = [{'id': supplier.id, 'text': supplier.name} for supplier in suppliers]
    return JsonResponse({'results': supplier_list})

def logistic_requirement_order_approved_list(request):
    # Obtener los ítems cuyas órdenes de requerimiento están aprobadas y el estado del ítem es Pendiente
    requirement_order_items = RequirementOrderItem.objects.filter(
        requirement_order__state='APROBADO',
        estado='P'  # Filtro adicional para solo obtener los ítems en estado Pendiente
    ).select_related(
        'sales_order_item', 
        'sales_order_item__salesorder',
        'sales_order_item__salesorder__project',
        'supplier'
    ).order_by('-requirement_order__created_at')

    # Obtener la lista de proveedores
    suppliers = Suppliers.objects.all()

    # Pasar los ítems aprobados y pendientes, y los proveedores al contexto para ser utilizados en el template
    context = {
        'requirement_order_items': requirement_order_items,
        'suppliers': suppliers,
    }

    return render(request, 'logistic_requirements/logistic_list.html', context)

@require_POST
def logistic_update_approved_items(request):
    updated_items = []
    errors = []

    # Iterar sobre los datos POST
    for key, value in request.POST.items():
        try:
            # Validar que la clave tenga al menos 2 partes tras el split y que el segundo elemento sea un número
            parts = key.split('_')
            if len(parts) == 2 and parts[1].isdigit():
                # Obtener el ID del ítem desde el nombre del campo
                item_id = int(parts[1])

                # Obtener el ítem correspondiente de la base de datos
                try:
                    item = RequirementOrderItem.objects.get(id=item_id)
                except RequirementOrderItem.DoesNotExist:
                    errors.append(f"El ítem con ID {item_id} no existe.")
                    continue

                # Actualizar los campos según la clave
                if key.startswith('quantity_'):
                    try:
                        item.quantity_requested = int(value) if value else 0
                    except ValueError:
                        errors.append(f"Error al convertir la cantidad para el ítem {item_id}: {value}")
                        continue

                elif key.startswith('price_'):
                    try:
                        item.price = float(value) if value else 0.0
                    except ValueError:
                        errors.append(f"Error al convertir el precio para el ítem {item_id}: {value}")
                        continue

                elif key.startswith('notes_'):
                    item.notes = value

                elif key.startswith('supplier_'):
                    item.supplier_id = value

                elif key.startswith('estado_'):
                    item.estado = value

                # Guardar el ítem actualizado
                item.save()
                updated_items.append(item)
            else:
                # Si no tiene la estructura adecuada, añade un error
                errors.append(f"ID no válido: {parts[1] if len(parts) > 1 else 'desconocido'}")
        
        except Exception as e:
            errors.append(f"Error procesando el ítem: {str(e)}")
            continue

    # Si hubo errores, devolver un mensaje con los errores
    if errors:
        return JsonResponse({'message': 'Hubo errores al actualizar los ítems.', 'errors': errors}, status=400)
    """_summary_

    Returns:
        _type_: _description_
    """
    # Si todo fue bien, devolver solo un mensaje de éxito
    return JsonResponse({'message': 'Ítems actualizados con éxito'}, status=200)

def logistic_requirement_order_detail_partial(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    items = requirement_order.items.all() 
    return render(request, 'requirements_approved/requirement_order_detail_partial.html', {
        'requirement_order': requirement_order,
        'items': items,  # Pasamos los ítems a la plantilla parcial
    })

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def logistic_export_order_to_excel(request, pk):
    # Obtener la orden específica
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    items = requirement_order.items.all().select_related(
        'sales_order_item', 'supplier'
    ).order_by('sales_order_item__description')

    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Orden {requirement_order.order_number}"

    # Encabezado con detalles de la orden
    ws['A1'] = "Detalles de la Orden de Requerimiento"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:K1')

    details = [
        f"ID Orden: {requirement_order.order_number}",
        f"Proyecto: {requirement_order.sales_order.project.name}",
        f"Cliente: {requirement_order.sales_order.project.client.legal_name}",
        f"Fecha Solicitada: {requirement_order.requested_date}",
        f"Fecha Creada: {requirement_order.created_at}",
        f"Pedido: {requirement_order.notes or 'Sin notas'}",
    ]

    for idx, detail in enumerate(details, start=2):
        ws[f"A{idx}"] = detail

    # Encabezados de la tabla
    headers = [
        "SAP","ITEM", "DETALLE", "UNIDAD", 
        "CANTIDAD", "HORAS", 
        "PROVEEDOR", "ESTADO"
    ]
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar el ancho de las columnas
    column_widths = [15, 20, 25, 30, 10, 20, 15, 15, 25, 15, 10]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Agregar datos de los ítems
    for row_num, item in enumerate(items, start=11):
        sales_order_item = item.sales_order_item
        data = [
            sales_order_item.sap_code or "N/A",
            sales_order_item.description or "N/A",
            item.notes or "",
            sales_order_item.unit_of_measurement or "N/A",
            float(item.quantity_requested) if item.quantity_requested else 0,
            getattr(sales_order_item, "custom_quantity", "N/A"),
            item.supplier.name if item.supplier else "N/A",
            item.get_estado_display() or "N/A",
        ]
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num == 11:  # Estado
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="order_{requirement_order.order_number}_pending_items.xlsx"'
    wb.save(response)

    return response
