import re
from openpyxl import load_workbook
from django.shortcuts import get_object_or_404
from .models import SalesOrder, SalesOrderItem
import pdfplumber
from datetime import datetime
from decimal import Decimal, InvalidOperation
import logging

def limpiar_valor(valor):
    if isinstance(valor, str):
        valor = re.sub(r'[^\d.,]', '', valor)
        try:
            return Decimal(valor.replace(',', ''))
        except ValueError:
            return None
    return valor

def procesar_archivo_excel(archivo_excel, salesorder_id):
    workbook = load_workbook(archivo_excel)
    sheet = workbook.active

    # Verificar que la orden de venta existe
    salesorder = get_object_or_404(SalesOrder, pk=salesorder_id)

    # Nombres de las columnas
    col_nro_articulo = 'Número de artículo'
    col_desc_articulo = 'Descripción del artículo'
    col_cantidad = 'Cantidad'
    col_precio_bruto = 'Precio bruto'
    col_total_bruto = 'Total bruto (ML)'
    col_unidad_medida = 'Unidad de medida'  # Nueva columna para la unidad de medida (opcional)

    # Encuentra los índices de las columnas
    header = [cell.value for cell in sheet[1]]
    
    # Manejar la columna de unidad de medida de forma opcional
    idx_unidad_medida = None
    if col_unidad_medida in header:
        idx_unidad_medida = header.index(col_unidad_medida)

    idx_nro_articulo = header.index(col_nro_articulo)
    idx_desc_articulo = header.index(col_desc_articulo)
    idx_cantidad = header.index(col_cantidad)
    idx_precio_bruto = header.index(col_precio_bruto)
    idx_total_bruto = header.index(col_total_bruto)

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[idx_nro_articulo] is None or row[idx_desc_articulo] is None or row[idx_cantidad] is None:
            continue

        cantidad = limpiar_valor(row[idx_cantidad])
        precio_bruto = limpiar_valor(row[idx_precio_bruto])
        total_bruto = limpiar_valor(row[idx_total_bruto])

        # Obtener unidad de medida solo si la columna está presente
        unidad_medida = row[idx_unidad_medida] if idx_unidad_medida is not None else 'N/A'

        if cantidad is None:
            continue  # Si no se puede convertir la cantidad, saltar esta fila

        SalesOrderItem.objects.create(
            salesorder=salesorder,
            sap_code=row[idx_nro_articulo],
            description=row[idx_desc_articulo],
            amount=cantidad,
            price=precio_bruto,
            price_total=total_bruto,
            unit_of_measurement=unidad_medida  # Guardar la unidad de medida si existe, o 'N/A' si no
        )
        
def extraer_datos_pdf(pdf_file):
    """Extrae datos de un archivo PDF y retorna un diccionario con la información"""
    texto_completo = ""
    with pdfplumber.open(pdf_file) as pdf:
        for pagina in pdf.pages:
            texto_completo += pagina.extract_text() + "\n"

    # Expresiones regulares para extraer datos específicos
    patron_serie_correlativo = r"E\d{3}-\d+"
    resultado_serie_correlativo = re.search(patron_serie_correlativo, texto_completo)
    serie_correlativo = resultado_serie_correlativo.group() if resultado_serie_correlativo else None

    patron_fecha_emision = r"Fecha de Emisión\s*:\s*(\d{2}/\d{2}/\d{4})"
    resultado_fecha_emision = re.search(patron_fecha_emision, texto_completo)
    fecha_emision = datetime.strptime(resultado_fecha_emision.group(1), "%d/%m/%Y").strftime('%Y-%m-%d') if resultado_fecha_emision else None

    patron_cliente = r"Señor\(es\) :\s*(.*)\n"
    resultado_cliente = re.search(patron_cliente, texto_completo)
    cliente = resultado_cliente.group(1) if resultado_cliente else None

    patron_ruc_cliente = r"RUC :\s*(\d+)"
    resultado_ruc_cliente = re.search(patron_ruc_cliente, texto_completo)
    ruc_cliente = resultado_ruc_cliente.group(1) if resultado_ruc_cliente else None

    patron_tipo_moneda = r"Tipo de Moneda\s*:\s*(.*)\n"
    resultado_tipo_moneda = re.search(patron_tipo_moneda, texto_completo)
    tipo_moneda = resultado_tipo_moneda.group(1) if resultado_tipo_moneda else None

    patron_descripcion = r"Descripción\s*(.*?)(?=Sub Total Ventas|$)"
    resultado_descripcion = re.search(patron_descripcion, texto_completo, re.DOTALL)
    descripcion = resultado_descripcion.group(1).strip() if resultado_descripcion else None

    patron_importe_total = r"Importe Total\s*:\s*(\$|S/)(.*)\n"
    resultado_importe_total = re.search(patron_importe_total, texto_completo)
    importe_total = convertir_decimal(resultado_importe_total.group(2).replace(',', '')) if resultado_importe_total else None

    patron_detraccion = r"Monto detracción:\s*(\$|S/)(.*)\n"
    resultado_detraccion = re.search(patron_detraccion, texto_completo)
    detraccion = convertir_decimal(resultado_detraccion.group(2).replace(',', '')) if resultado_detraccion else None

    patron_neto_cobrar = r"Monto neto pendiente de pago\s*:\s*(\$|S/)(.*)\n"
    resultado_neto_cobrar = re.search(patron_neto_cobrar, texto_completo)
    neto_cobrar = convertir_decimal(resultado_neto_cobrar.group(2).replace(',', '')) if resultado_neto_cobrar else None

    patron_total_cuotas = r"Total de Cuotas\s*:\s*(\d+)"
    resultado_total_cuotas = re.search(patron_total_cuotas, texto_completo)
    total_cuotas = int(resultado_total_cuotas.group(1)) if resultado_total_cuotas else None

    patron_fecha_vencimiento = r"Fec. Venc.\s*Monto\n\d+\s*(\d{2}/\d{2}/\d{4})"
    resultado_fecha_vencimiento = re.search(patron_fecha_vencimiento, texto_completo)
    fecha_vencimiento = datetime.strptime(resultado_fecha_vencimiento.group(1), "%d/%m/%Y").strftime('%Y-%m-%d') if resultado_fecha_vencimiento else None

    return {
        "serie_correlativo": serie_correlativo,
        "fecha_emision": fecha_emision,
        "cliente": cliente,
        "ruc_cliente": ruc_cliente,
        "tipo_moneda": tipo_moneda,
        "descripcion": descripcion,
        "importe_total": importe_total,
        "detraccion": detraccion,
        "monto_neto_cobrar": neto_cobrar,
        "total_cuotas": total_cuotas,
        "fecha_vencimiento": fecha_vencimiento,
    }

def convertir_decimal(cadena):
    try:
        return Decimal(cadena.strip()) if cadena else None
    except InvalidOperation:
        logging.error(f"Error al convertir {cadena} a Decimal")
        return None
