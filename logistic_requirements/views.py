from datetime import datetime
from django.shortcuts import render, get_object_or_404
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db import transaction
from django.db.models import Q
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from logistic_inventory.models import Item
from logistic_suppliers.models import Suppliers
from accounting_order_sales.models import PurchaseOrder, PurchaseOrderItem
from .forms import RequirementOrderForm, RequirementOrderItemFormSet , RequirementOrderListForm
from .models import RequirementOrder, RequirementOrderItem
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import RequirementOrderItem
from django.views.decorators.csrf import csrf_exempt
from django.utils.timezone import localdate
import json

# Vista para listar todas las RequirementOrders aprobadas con ítems en estado Pendiente o todas las órdenes sin filtros
class RequirementOrderListView(ListView):
    model = RequirementOrder
    template_name = 'requirement_order_index.html'
    context_object_name = 'requirement_orders'

    def get_queryset(self):
        # Obtener los parámetros de los filtros
        show_pending = self.request.GET.get('show_pending') == 'true'
        show_comprando = self.request.GET.get('show_comprando') == 'true'
        show_all = self.request.GET.get('show_all') == 'true'
        
        # Iniciar queryset con todas las órdenes APROBADAS por contabilidad
        queryset = RequirementOrder.objects.order_by('-id').prefetch_related('items') # filter(state='APROBADO')
        
        if show_all:
            queryset = queryset.distinct()
            print(show_all)
        elif show_comprando:
            # Mostrar solo las órdenes con ítems en estado Pendiente o Comprando
            queryset = queryset.filter(state='APROBADO').filter(
                    items__estado='C'# items__estado__in=['P', 'C']
            ).distinct()
            print(show_comprando)
        elif show_pending:
            # Mostrar solo las órdenes con ítems en estado Pendiente o Comprando
            queryset = queryset.filter(state='APROBADO').filter(
                    items__estado__in=['P', 'C']
            ).distinct()
            print(show_pending)
        else:
            queryset = queryset.filter(state='APROBADO').filter(
                items__estado='P' #items__estado__in=['P', 'C']
            ).distinct()
         
        
        # Calcular el estado general de cada orden
        for order in queryset:
            items = order.items.all()
            total_items = items.count()

            if total_items == 0:
                order.global_state = "No tiene ítems"
                continue

            ready_count = items.filter(estado='L').count()
            buying_count = items.filter(estado='C').count()
            pending_count = items.filter(estado='P').count()

            # Determinar el estado general
            if ready_count == total_items:
                order.global_state = "Listo"
            elif buying_count >= total_items / 2:
                order.global_state = "Comprando"
            elif pending_count > 0:
                order.global_state = "Pendiente"
            else:
                order.global_state = "Completado"  # Si no hay items pendientes, listos o comprando

        return queryset


def search_requirement_order_list_view(request):
    query = request.GET.get('q', '')
    print(f"Search Query: {query}")  # Agregar log

    if query:
        requirement_orders = RequirementOrder.objects.filter(Q(order_number__icontains=query) | Q(estado__icontains=query)
                                                            | Q(notes__icontains=query)).order_by('-id')
    else:
        requirement_orders = RequirementOrder.objects.all().order_by('-id')
    
    print(requirement_orders)

    context = {'requirement_orders': requirement_orders, 'form': RequirementOrderListForm()}
    return render(request, 'requirement_order_list.html', context)

def requirement_order_detail_view(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    filtrar = request.GET.get('filtrar')

    # Filtrar los ítems en estado 'P' si se requiere, o cargar todos los ítems ordenados por categoría
    if filtrar == 'P':
        items = requirement_order.items.filter(estado='P').select_related('sales_order_item').order_by('sales_order_item__category')
    else:
        items = requirement_order.items.all().select_related('sales_order_item').order_by('sales_order_item__category')

    suppliers = Suppliers.objects.all()

    # Crear un diccionario de inventario disponible, usando sap como clave
    inventory_data = {item.item.sap: item.quantity for item in Item.objects.select_related('item').all()}

    # Agregar disponibilidad y tiempo de servicio a cada item en el queryset
    for item in items:
        item.disponible_inventario = inventory_data.get(item.sap_code, 0)

        # Calcular tiempo de servicio en horas solo si la categoría no es 'Material', 'Consumible', o 'Equipo'
        if item.sales_order_item.category not in ["Material", "Consumible", "Equipo"]:
            item.tiempo_servicio = requirement_order.sales_order.days * 8 * item.quantity_requested
        else:
            item.tiempo_servicio = None

    return render(request, 'requirement_order_detail.html', {
        'requirement_order': requirement_order,
        'items': items,
        'suppliers': suppliers,
        'filtrar': filtrar,
    })

@require_POST
def update_requirement_order_items(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    updated_items = []

    # Procesar cada ítem en la orden de requerimiento
    for item in requirement_order.items.all():
        item.quantity_requested = request.POST.get(f'quantity_requested_{item.id}', item.quantity_requested)
        item.price = request.POST.get(f'price_{item.id}', item.price)
        item.notes = request.POST.get(f'notes_{item.id}', item.notes)
        item.supplier_id = request.POST.get(f'supplier_{item.id}')
        item.estado = request.POST.get(f'estado_{item.id}', item.estado)

        # Guardar la fecha de la orden de compra
        date_str = request.POST.get(f'date_{item.id}')
        if date_str:
            try:
                item.date_purchase_order = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                item.date_purchase_order = None


        updated_items.append(item)

    # Usar bulk_update para mejorar el rendimiento
    RequirementOrderItem.objects.bulk_update(
        updated_items, 
        ['quantity_requested', 'price', 'notes', 'supplier_id', 'estado', 'date_purchase_order']
    )

    # Recalcular remaining_requirement para todos los sales_order_items relacionados
    for item in updated_items:
        item.sales_order_item.update_remaining_requirement()
        print(f'Item ID: {item.id}, Fecha: {item.date_purchase_order}')


    # Retornar el mensaje directamente en HTML
    return HttpResponse(
        '<div class="alert alert-success" role="alert">Items actualizados con éxito</div>',
        status=200
    )



def create_purchase_order(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)

    # Verificar si ya existe una orden de compra para esta orden de requerimiento
    if requirement_order.purchase_order_created:
        error_message = "<div>Ya se se ha creado una Orden de Compra para esta Orden de Requerimiento.</div>"
        return HttpResponse(error_message, content_type="text/html")

    # Filtrar los ítems que están en estado "C"
    items_comprando = RequirementOrderItem.objects.filter(requirement_order=requirement_order, estado='C')
    
    if not items_comprando.exists():
        error_message = "<div>No hay ítems en estado 'Comprando' para crear una Orden de Compra.</div>"
        return HttpResponse(error_message, content_type="text/html")

    # Crear la PurchaseOrder
    with transaction.atomic():
        purchase_order = PurchaseOrder.objects.create(
            salesorder=requirement_order.sales_order,
            description=f"{requirement_order.notes} - {requirement_order.order_number}",
            requested_date=requirement_order.requested_date,
            requested_by=request.user.username if request.user else 'Desconocido',
            acepted=True
        )

        # Crear los PurchaseOrderItems asociados a los ítems comprando
        purchase_order_items = [
            PurchaseOrderItem(
                purchaseorder=purchase_order,
                sales_order_item=item.sales_order_item,
                sap_code=item.sap_code,
                quantity_requested=item.quantity_requested,
                price=item.price,
                price_total=item.total_price,
                notes=item.notes,
                supplier=item.supplier
            )
            for item in items_comprando
        ]
        PurchaseOrderItem.objects.bulk_create(purchase_order_items)

        # Actualizar la RequirementOrder para indicar que la orden de compra ha sido creada
        requirement_order.purchase_order_created = True
        requirement_order.save()

    # Respuesta de éxito en HTML
    success_message = f"<div>Orden de Compra creada para la Orden de Requerimiento #{requirement_order.order_number}.</div>"
    return HttpResponse(success_message, content_type="text/html")

def ajax_load_suppliers(request):
    term = request.GET.get('term', '')
    suppliers = Suppliers.objects.filter(name__icontains=term)[:20]
    supplier_list = [{'id': supplier.id, 'text': supplier.name} for supplier in suppliers]
    return JsonResponse({'results': supplier_list})

def requirement_order_approved_list(request):
    # Obtener los ítems cuyas órdenes de requerimiento están aprobadas y el estado del ítem es Pendiente
    requirement_order_items = RequirementOrderItem.objects.filter(
        requirement_order__state='APROBADO',
        estado='P'  # Filtro adicional para solo obtener los ítems en estado Pendiente
    ).select_related(
        'sales_order_item', 
        'sales_order_item__salesorder',
        'sales_order_item__salesorder__project',
        'supplier'
    ).order_by('-requirement_order__created_at')

    # Obtener la lista de proveedores
    suppliers = Suppliers.objects.all()

    # Pasar los ítems aprobados y pendientes, y los proveedores al contexto para ser utilizados en el template
    context = {
        'requirement_order_items': requirement_order_items,
        'suppliers': suppliers,
    }

    return render(request, 'requirements_approved/requirement_order_approved_list.html', context)

@require_POST
def update_approved_items(request):
    updated_items = []
    errors = []

    # Iterar sobre los datos POST
    for key, value in request.POST.items():
        try:
            # Validar que la clave tenga al menos 2 partes tras el split y que el segundo elemento sea un número
            parts = key.split('_')
            if len(parts) == 2 and parts[1].isdigit():
                # Obtener el ID del ítem desde el nombre del campo
                item_id = int(parts[1])

                # Obtener el ítem correspondiente de la base de datos
                try:
                    item = RequirementOrderItem.objects.get(id=item_id)
                except RequirementOrderItem.DoesNotExist:
                    errors.append(f"El ítem con ID {item_id} no existe.")
                    continue

                # Actualizar los campos según la clave
                if key.startswith('quantity_'):
                    try:
                        item.quantity_requested = int(value) if value else 0
                    except ValueError:
                        errors.append(f"Error al convertir la cantidad para el ítem {item_id}: {value}")
                        continue

                elif key.startswith('price_'):
                    try:
                        item.price = float(value) if value else 0.0
                    except ValueError:
                        errors.append(f"Error al convertir el precio para el ítem {item_id}: {value}")
                        continue

                elif key.startswith('notes_'):
                    item.notes = value

                elif key.startswith('supplier_'):
                    item.supplier_id = value

                elif key.startswith('estado_'):
                    item.estado = value

                # Guardar el ítem actualizado
                item.save()
                updated_items.append(item)
            else:
                # Si no tiene la estructura adecuada, añade un error
                errors.append(f"ID no válido: {parts[1] if len(parts) > 1 else 'desconocido'}")
        
        except Exception as e:
            errors.append(f"Error procesando el ítem: {str(e)}")
            continue

    # Si hubo errores, devolver un mensaje con los errores
    if errors:
        return JsonResponse({'message': 'Hubo errores al actualizar los ítems.', 'errors': errors}, status=400)
    """_summary_

    Returns:
        _type_: _description_
    """
    # Si todo fue bien, devolver solo un mensaje de éxito
    return JsonResponse({'message': 'Ítems actualizados con éxito'}, status=200)

def requirement_order_detail_partial(request, pk):
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    items = requirement_order.items.all() 
    return render(request, 'requirements_approved/requirement_order_detail_partial.html', {
        'requirement_order': requirement_order,
        'items': items,  # Pasamos los ítems a la plantilla parcial
    })

def export_order_to_excel(request, pk):
    # Obtener la orden específica
    requirement_order = get_object_or_404(RequirementOrder, pk=pk)
    items = requirement_order.items.filter(estado='P')

    # Crear un nuevo libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Orden {requirement_order.order_number}"

    # Encabezado con detalles de la orden
    ws['A1'] = "Detalles de la Orden de Requerimiento"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:K1')

    details = [
        f"ID Orden: {requirement_order.order_number}",
        f"Proyecto: {requirement_order.sales_order.project.name}",
        f"Cliente: {requirement_order.sales_order.project.client.legal_name}",
        f"Fecha Solicitada: {requirement_order.requested_date}",
        f"Fecha Creada: {requirement_order.created_at}",
        f"Notas: {requirement_order.notes}",
    ]

    for idx, detail in enumerate(details, start=2):
        ws[f"A{idx}"] = detail

    # Encabezados de la tabla
    headers = [
        "SAP", "Categoría", "Ítem", "Detalle", "Unidad", 
        "Cantidad en Unidades", "Cantidad en Horas", 
        "Stock", "Proveedor", "Documento", "Estado"
    ]
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col_num, value=column_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Ajustar el ancho de las columnas
    column_widths = [15, 20, 25, 30, 10, 20, 15, 15, 25, 15, 10]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Agregar datos de los ítems
    for row_num, item in enumerate(items, start=11):
        sales_order_item = item.sales_order_item
        data = [
            sales_order_item.sap_code,
            sales_order_item.category,
            sales_order_item.description,
            item.notes or "",
            sales_order_item.unit_of_measurement,
            item.quantity_requested,
            getattr(sales_order_item, "custom_quantity", "N/A"),
            item.sales_order_item.remaining_requirement,
            item.supplier.name if item.supplier else "N/A",
            "Sí" if item.file_attachment else "No",
            item.get_estado_display(),
        ]
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num == 11:  # Estado
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Configurar la respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="order_{requirement_order.order_number}_pending_items.xlsx"'
    wb.save(response)

    return response


#caja chica
def logistic_petty_cash(request):
    # Obtener la fecha de hoy según la zona horaria configurada
    today = localdate()

    # Obtenemos los parámetros del rango de fechas (si existen)
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Si no se han proporcionado fechas, mostrar ítems del día actual
    if not start_date and not end_date:
        items = PurchaseOrderItem.objects.filter(purchaseorder__scheduled_date=today).select_related(
            'purchaseorder', 'sales_order_item__salesorder', 'supplier'
        )
    else:
        # Si se proporcionan fechas, buscar entre esas dos fechas
        if not end_date:
            end_date = today  # Si solo hay fecha de inicio, el rango termina en el día actual

        items = PurchaseOrderItem.objects.filter(
            purchaseorder__scheduled_date__range=[start_date, end_date]
        ).select_related('purchaseorder', 'sales_order_item__salesorder', 'supplier')

    context = {
        'items': items,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'logistic_pettycash/petty_cash_items.html', context)

def logistic_petty_cash_state(request):
    # Obtener los parámetros de filtro
    payment_status = request.GET.get('status')  # Puede ser "Pagado" o "No Pagado"
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Inicializar el queryset base
    items = PurchaseOrderItem.objects.select_related(
        'purchaseorder', 'sales_order_item__salesorder', 'supplier'
    )

    # Filtrar por estado de pago si está definido
    if payment_status:
        items = items.filter(payment_status=payment_status)

    # Si no se filtra por estado de pago, aplicar el filtro de fechas
    elif start_date or end_date:
        if not end_date:
            end_date = localdate()  # Fecha actual como final si no se especifica
        if not start_date:
            start_date = localdate()  # Fecha actual como inicio si no se especifica

        items = items.filter(purchaseorder__scheduled_date__range=[start_date, end_date])

    context = {
        'items': items,
        'start_date': start_date,
        'end_date': end_date,
        'payment_status': payment_status,
        'class_pay_choices': PurchaseOrderItem.CLASS_PAY_CHOICES,
        'type_pay_choices': PurchaseOrderItem.TYPE_PAY_CHOICES,
    }
    return render(request, 'logistic_pettycash/petty_cash_state.html', context)

def update_payment_status(request, item_id):
    item = get_object_or_404(PurchaseOrderItem, id=item_id)
    item.payment_status = 'Pagado' if item.payment_status == 'No Pagado' else 'No Pagado'
    item.save()
    return JsonResponse({'status': item.payment_status})

@csrf_exempt
def update_field(request, item_id):
    try:
        if request.method == 'POST':
            data = json.loads(request.body)
            field = data.get('field')
            value = data.get('value')

            item = get_object_or_404(PurchaseOrderItem, id=item_id)

            if field in ['class_pay', 'type_pay']:
                setattr(item, field, value)
                item.save()
                return JsonResponse({'success': True, 'field': field, 'value': value})

            return JsonResponse({'success': False, 'error': 'Campo no válido.'}, status=400)

        return JsonResponse({'success': False, 'error': 'Método no permitido.'}, status=405)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)
